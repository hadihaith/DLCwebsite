
from .models import Application, ExchangeApplication, ExchangeNomination, PartnerUniversity, User, DeanListStudent, DeanList, Thread, Reply, ThreadSettings, majors
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseForbidden
from django.shortcuts import render
from django.urls import reverse
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.base import ContentFile
from datetime import datetime, timedelta
from .utils import process_dean_list_excel
from user_agents import parse
from django.contrib import messages
from django.db.models import Q, Count
import random
from decimal import Decimal, InvalidOperation
import os
import re
from .forms import EventForm, ExchangeApplicationForm, PartnerUniversityForm
import requests
from django.http import HttpResponse, HttpResponseBadRequest
from urllib.parse import urlparse, quote
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404


def cleanup_invalid_course_ids():
    """
    Utility function to delete courses with invalid course IDs.
    Valid course IDs should have letters followed by exactly 3 digits (e.g., ACC201, FIN305).
    Invalid examples: ACC0, MGTMKT1011220, ACC, FIN12, etc.
    """
    try:
        from .models import Course
        
        # Pattern to match valid course IDs: letters followed by exactly 3 digits
        valid_pattern = r'^[A-Z]+\d{3}$'
        
        # Find all courses with invalid IDs
        all_courses = Course.objects.all()
        invalid_courses = []
        
        for course in all_courses:
            if not re.match(valid_pattern, course.course_id):
                invalid_courses.append(course)
        
        if invalid_courses:
            print(f"Found {len(invalid_courses)} courses with invalid course IDs:")
            for course in invalid_courses:
                print(f"  - {course.course_id}: {course.course_name}")
            
            # Delete invalid courses
            invalid_count = len(invalid_courses)
            Course.objects.filter(id__in=[c.id for c in invalid_courses]).delete()
            
            print(f"Successfully deleted {invalid_count} courses with invalid course IDs")
            return invalid_count
        else:
            print("No courses with invalid course IDs found.")
            return 0
            
    except Exception as e:
        print(f"Error during course cleanup: {str(e)}")
        return -1
from django.db.models import Q

def home(request):
    return render(request, 'frontend/index.html')

def resources(request):
    """Resources page with three main sections"""
    return render(request, 'frontend/resources.html')

def student_guide(request):
    """Student guide page with downloadable guides by year"""
    selected_year = request.GET.get('year', '2023')  # Default to 2023
    
    # Student guide data organized by year
    student_guides = {
        '2023': [
            {
                'title': 'Information Systems Management (نظم المعلومات الادارية)',
                'description': 'Study guide for Information Systems Management major 2023',
                'download_url': 'https://drive.google.com/file/d/1lakt3HdGxPv4cVSxvgOEMg6D25i2L3Zl/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1lakt3HdGxPv4cVSxvgOEMg6D25i2L3Zl/preview',
                'category': 'Information Systems',
                'size': 'PDF'
            },
            {
                'title': 'Public Administration (الادارة العامه)',
                'description': 'Study guide for Public Administration major 2023',
                'download_url': 'https://drive.google.com/file/d/1I675DMORINOxrUENdvo8FvKJNO2na5OS/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1I675DMORINOxrUENdvo8FvKJNO2na5OS/preview',
                'category': 'Public Administration',
                'size': 'PDF'
            },
            {
                'title': 'Economics (الاقتصاد)',
                'description': 'Study guide for Economics major 2023',
                'download_url': 'https://drive.google.com/file/d/1x19PYd7PMwMOMJIpb4kCfjucRHGzmQ14/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1x19PYd7PMwMOMJIpb4kCfjucRHGzmQ14/preview',
                'category': 'Economics',
                'size': 'PDF'
            },
            {
                'title': 'Marketing (التسويق)',
                'description': 'Study guide for Marketing major 2023',
                'download_url': 'https://drive.google.com/file/d/1iTukxKgAfGcepcYi4yHpIrFEdjGOuFZn/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1iTukxKgAfGcepcYi4yHpIrFEdjGOuFZn/preview',
                'category': 'Marketing',
                'size': 'PDF'
            },
            {
                'title': 'Management (الادارة)',
                'description': 'Study guide for Management major 2023',
                'download_url': 'https://drive.google.com/file/d/198SS0XmI0VsJUf_2sisrCkNfEuKFeGw_/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/198SS0XmI0VsJUf_2sisrCkNfEuKFeGw_/preview',
                'category': 'Management',
                'size': 'PDF'
            },
            {
                'title': 'Accounting (المحاسبة)',
                'description': 'Study guide for Accounting major 2023',
                'download_url': 'https://drive.google.com/file/d/1VynuiD99Oo0Hm51RE8Gwnjj7MYVxeq0w/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1VynuiD99Oo0Hm51RE8Gwnjj7MYVxeq0w/preview',
                'category': 'Accounting',
                'size': 'PDF'
            },
            {
                'title': 'Operations Management & Supply Chain (ادارة العمليات وسلسلة الامدادات)',
                'description': 'Study guide for Operations Management & Supply Chain major 2023',
                'download_url': 'https://drive.google.com/file/d/1DtKEq9C0J3Gv2KNbQPSaZYLzsFpYAxCG/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1DtKEq9C0J3Gv2KNbQPSaZYLzsFpYAxCG/preview',
                'category': 'Operations Management',
                'size': 'PDF'
            },
            {
                'title': 'Finance & Financial Institutions (التمويل و المنشآت الماليه)',
                'description': 'Study guide for Finance & Financial Institutions major 2023',
                'download_url': 'https://drive.google.com/file/d/1ohcceBnBC48cHOfA-YLMRETWC7sVpRqb/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1ohcceBnBC48cHOfA-YLMRETWC7sVpRqb/preview',
                'category': 'Finance',
                'size': 'PDF'
            }
        ],
        '2024': [
            {
                'title': 'Management (الادارة)',
                'description': 'Study guide for Management major 2024',
                'download_url': 'https://drive.google.com/file/d/1OX-YpB0ZaX1uH8NoKCCGTOHR7EErXFg2/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1OX-YpB0ZaX1uH8NoKCCGTOHR7EErXFg2/preview',
                'category': 'Management',
                'size': 'PDF'
            },
            {
                'title': 'Public Administration (الادارة العامه)',
                'description': 'Study guide for Public Administration major 2024',
                'download_url': 'https://drive.google.com/file/d/11QoBKNF3iJz25rRKExSaGFDpDQmBL5Pw/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/11QoBKNF3iJz25rRKExSaGFDpDQmBL5Pw/preview',
                'category': 'Public Administration',
                'size': 'PDF'
            },
            {
                'title': 'Accounting (المحاسبة)',
                'description': 'Study guide for Accounting major 2024',
                'download_url': 'https://drive.google.com/file/d/1m34Rpp1cDsVgLPqDk1h9ys66ChQ3WxYx/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1m34Rpp1cDsVgLPqDk1h9ys66ChQ3WxYx/preview',
                'category': 'Accounting',
                'size': 'PDF'
            },
            {
                'title': 'Economics (الاقتصاد)',
                'description': 'Study guide for Economics major 2024',
                'download_url': 'https://drive.google.com/file/d/1ydN5ZCdIA8XXwEZDC-_sikMZYU2PcFr8/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1ydN5ZCdIA8XXwEZDC-_sikMZYU2PcFr8/preview',
                'category': 'Economics',
                'size': 'PDF'
            },
            {
                'title': 'Marketing (التسويق)',
                'description': 'Study guide for Marketing major 2024',
                'download_url': 'https://drive.google.com/file/d/1fj9E_6rbGzgfFnRBdhuVx29aLdzBz5ba/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1fj9E_6rbGzgfFnRBdhuVx29aLdzBz5ba/preview',
                'category': 'Marketing',
                'size': 'PDF'
            },
            {
                'title': 'Operations Management & Supply Chain (ادارة العمليات وسلسلة الامدادات)',
                'description': 'Study guide for Operations Management & Supply Chain major 2024',
                'download_url': 'https://drive.google.com/file/d/1P_Xad-CX_E1ztf9EkZ_mI2la_OXyOic5/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1P_Xad-CX_E1ztf9EkZ_mI2la_OXyOic5/preview',
                'category': 'Operations Management',
                'size': 'PDF'
            },
            {
                'title': 'Finance & Financial Institutions (التمويل والمنشآت المالية)',
                'description': 'Study guide for Finance & Financial Institutions major 2024',
                'download_url': 'https://drive.google.com/file/d/1wtArpekA6YVFJpCiMurdUgk-q1Tv0oqD/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1wtArpekA6YVFJpCiMurdUgk-q1Tv0oqD/preview',
                'category': 'Finance',
                'size': 'PDF'
            }
        ],
        '2025': [
            {
                'title': 'Public Administration (الادارة العامه)',
                'description': 'Study guide for Public Administration major 2025',
                'download_url': 'https://drive.google.com/file/d/1W_D0FPwy9a08xtg5_amD__qBAusxjcZD/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1W_D0FPwy9a08xtg5_amD__qBAusxjcZD/preview',
                'category': 'Public Administration',
                'size': 'PDF'
            },
            {
                'title': 'Economics (الاقتصاد)',
                'description': 'Study guide for Economics major 2025',
                'download_url': 'https://drive.google.com/file/d/1uCxk6tg1DpTBhzWG4Dej5vGbo__SBmPZ/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1uCxk6tg1DpTBhzWG4Dej5vGbo__SBmPZ/preview',
                'category': 'Economics',
                'size': 'PDF'
            },
            {
                'title': 'Marketing (التسويق)',
                'description': 'Study guide for Marketing major 2025',
                'download_url': 'https://drive.google.com/file/d/1_emotOjnlXXxMMOCZXkuniinLct2SQJs/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1_emotOjnlXXxMMOCZXkuniinLct2SQJs/preview',
                'category': 'Marketing',
                'size': 'PDF'
            },
            {
                'title': 'Operations Management & Supply Chain (ادارة العمليات وسلسلة الامدادات)',
                'description': 'Study guide for Operations Management & Supply Chain major 2025',
                'download_url': 'https://drive.google.com/file/d/1u8e3ovjp-8KXM1m9cwER--U0WW2iGICj/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1u8e3ovjp-8KXM1m9cwER--U0WW2iGICj/preview',
                'category': 'Operations Management',
                'size': 'PDF'
            },
            {
                'title': 'Information Systems Management (نظم المعلومات)',
                'description': 'Study guide for Information Systems Management major 2025',
                'download_url': 'https://drive.google.com/file/d/1KtUsdcfd5kHnDvVrvj3S54sG_DRiekYu/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1KtUsdcfd5kHnDvVrvj3S54sG_DRiekYu/preview',
                'category': 'Information Systems',
                'size': 'PDF'
            },
            {
                'title': 'Finance & Financial Institutions (التمويل والمنشآت المالية)',
                'description': 'Study guide for Finance & Financial Institutions major 2025',
                'download_url': 'https://drive.google.com/file/d/1_UmQ2BwcrABtgYOIG-qVKmJxD0S099yA/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1_UmQ2BwcrABtgYOIG-qVKmJxD0S099yA/preview',
                'category': 'Finance',
                'size': 'PDF'
            },
            {
                'title': 'Accounting (المحاسبة)',
                'description': 'Study guide for Accounting major 2025',
                'download_url': 'https://drive.google.com/file/d/1tqPPjTgJyxk_pyhx5SWHfwWTNnuSsf0g/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1tqPPjTgJyxk_pyhx5SWHfwWTNnuSsf0g/preview',
                'category': 'Accounting',
                'size': 'PDF'
            },
            {
                'title': 'Management (الادارة)',
                'description': 'Study guide for Management major 2025',
                'download_url': 'https://drive.google.com/file/d/1iiJQGxDt609ZvtdyVS7qeRgFYQX9smfq/view?usp=drive_link',
                'view_url': 'https://drive.google.com/file/d/1iiJQGxDt609ZvtdyVS7qeRgFYQX9smfq/preview',
                'category': 'Management',
                'size': 'PDF'
            }
        ]
    }
    
    # Get guides for selected year
    guides = student_guides.get(selected_year, [])
    
    # Group guides by category
    guides_by_category = {}
    for guide in guides:
        category = guide['category']
        if category not in guides_by_category:
            guides_by_category[category] = []
        guides_by_category[category].append(guide)
    
    context = {
        'selected_year': selected_year,
        'available_years': ['2023', '2024', '2025'],
        'guides': guides,
        'guides_by_category': guides_by_category,
        'total_guides': len(guides)
    }
    
    return render(request, 'frontend/student_guide.html', context)

def course_descriptions(request):
    """Course descriptions and syllabuses page"""
    try:
        from .models import Course
        from django.core.paginator import Paginator
        
        # Get all courses ordered by course ID
        courses = Course.objects.filter(is_active=True).order_by('course_id')
        
        # Filter by department if requested
        department = request.GET.get('department')
        if department:
            courses = courses.filter(department=department)
        
        # Search functionality
        search_query = request.GET.get('search')
        if search_query:
            courses = courses.filter(
                Q(course_id__icontains=search_query) | 
                Q(course_name__icontains=search_query)
            )
        
        # Pagination
        paginator = Paginator(courses, 25)  # Show 25 courses per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Get unique departments for filter dropdown
        departments_queryset = Course.objects.filter(is_active=True).values_list('department', flat=True)
        departments = list(set([dept for dept in departments_queryset if dept]))  # Use set() to ensure uniqueness
        
        context = {
            'page_obj': page_obj,
            'courses': page_obj,
            'departments': sorted(departments),
            'current_department': department,
            'search_query': search_query,
            'total_courses': courses.count(),
        }
        
    except Exception as e:
        # Handle case where Course model doesn't exist or database issues
        context = {
            'page_obj': None,
            'courses': [],
            'departments': [],
            'current_department': None,
            'search_query': request.GET.get('search'),
            'total_courses': 0,
            'error_message': 'Course data is not available. Please contact the administrator.'
        }
    
    return render(request, 'frontend/course_descriptions.html', context)

def apply(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        student_id = request.POST.get('student_id')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        GPA = request.POST.get('GPA')
        passed_credits = request.POST.get('passed_credits')
        major = request.POST.get('major')
        anything_else = request.POST.get('anything_else')
        applications = Application.objects.all()
        if applications.filter(student_id=student_id).exists():
            return render(request, 'frontend/apply.html', {'error': 'You have already submitted an application. Please wait for a response before submitting another.'})
        application = Application(
            name=name,
            student_id=student_id,
            phone=phone,
            email=email,
            passed_credits=passed_credits,
            GPA=GPA,
            major=major,
            anything_else=anything_else
        )
        
        application.save()
        return render(request, 'frontend/success.html', {'application': application})
    return render(request, 'frontend/apply.html')


def parking_application(request):
    from .forms import ParkingApplicationForm
    from .models import ParkingApplication, DeanListStudent
    
    submitted = False
    error_message = None
    
    if request.method == 'POST':
        form = ParkingApplicationForm(request.POST)
        if form.is_valid():
            student_id = form.cleaned_data['student_id']
            
            # Check if student is in the latest dean's list
            latest_dean_list = DeanListStudent.objects.order_by('-year', '-semester').values('semester', 'year').first()
            
            if latest_dean_list:
                is_in_deans_list = DeanListStudent.objects.filter(
                    student_id=student_id,
                    semester=latest_dean_list['semester'],
                    year=latest_dean_list['year']
                ).exists()
                
                if not is_in_deans_list:
                    error_message = 'You must be in the latest published Dean\'s List to apply for parking.'
                else:
                    # Check if already applied
                    if ParkingApplication.objects.filter(student_id=student_id).exists():
                        error_message = 'You have already submitted a parking application.'
                    else:
                        application = form.save()
                        # Redirect to success page
                        return render(request, 'frontend/parking_success.html', {
                            'application': application
                        })
            else:
                error_message = 'No Dean\'s List data available. Please contact administration.'
    else:
        form = ParkingApplicationForm()
    
    context = {
        'form': form,
        'error_message': error_message,
    }
    return render(request, 'frontend/parking_application.html', context)


@login_required
def parking_management(request):
    """
    Manage parking spot applications - view all applications with eligibility verification
    """
    from .models import ParkingApplication, DeanListStudent
    from decimal import Decimal
    
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # Get latest Dean's List info for verification
    latest_dean_list = DeanListStudent.objects.order_by('-year', '-semester').values('semester', 'year').first()
    
    # Get all parking applications ordered by GPA (highest first)
    all_applications = ParkingApplication.objects.all()
    
    # Separate eligible and ineligible applications
    eligible_apps = []
    ineligible_apps = []
    
    for app in all_applications:
        # Check eligibility criteria
        gpa_check = app.gpa >= Decimal('3.5')
        license_check = app.has_kuwaiti_license
        
        # Check if in latest Dean's List
        deans_list_check = False
        if latest_dean_list:
            deans_list_check = DeanListStudent.objects.filter(
                student_id=app.student_id,
                semester=latest_dean_list['semester'],
                year=latest_dean_list['year']
            ).exists()
        
        is_eligible = gpa_check and license_check and deans_list_check
        
        if is_eligible:
            eligible_apps.append(app)
        else:
            ineligible_apps.append(app)
    
    # Eligible apps are already sorted by -gpa, -submitted_at in model Meta ordering
    # Sort ineligible apps the same way
    ineligible_apps.sort(key=lambda x: (-x.gpa, -x.submitted_at))
    
    # Combine lists with eligible first, then ineligible
    applications = eligible_apps + ineligible_apps
    
    # Add eligibility status and rejection reasons to each application
    applications_with_status = []
    separator_added = False
    
    for app in applications:
        # Check all eligibility criteria
        gpa_check = app.gpa >= Decimal('3.5')
        license_check = app.has_kuwaiti_license
        
        deans_list_check = False
        if latest_dean_list:
            deans_list_check = DeanListStudent.objects.filter(
                student_id=app.student_id,
                semester=latest_dean_list['semester'],
                year=latest_dean_list['year']
            ).exists()
        
        is_eligible = gpa_check and license_check and deans_list_check
        
        # Add separator before first ineligible application
        add_separator = not is_eligible and not separator_added and len(eligible_apps) > 0
        if add_separator:
            separator_added = True
        
        # Determine rejection reasons
        rejection_reasons = []
        if not gpa_check:
            rejection_reasons.append(f"GPA below 3.5 ({app.gpa})")
        if not license_check:
            rejection_reasons.append("No Kuwaiti driver's license")
        if not deans_list_check:
            rejection_reasons.append("Not in latest Dean's List")
        
        # Get major display name
        major_display = dict(majors).get(app.major, app.major)
        
        applications_with_status.append({
            'application': app,
            'eligible': is_eligible,
            'rejection_reasons': rejection_reasons,
            'major_display': major_display,
            'add_separator': add_separator
        })
    
    # Calculate statistics
    total_applications = len(applications_with_status)
    eligible_applications = len([app for app in applications_with_status if app['eligible']])
    ineligible_applications = total_applications - eligible_applications
    
    # Average GPA for eligible applications
    avg_gpa = 0
    if eligible_applications > 0:
        avg_gpa = sum([app['application'].gpa for app in applications_with_status if app['eligible']]) / eligible_applications
    
    # Check if user can delete applications
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    can_delete = request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles
    can_delete_all = can_delete
    
    context = {
        'applications_with_status': applications_with_status,
        'total_applications': total_applications,
        'eligible_applications': eligible_applications,
        'ineligible_applications': ineligible_applications,
        'avg_gpa': avg_gpa,
        'can_delete': can_delete,
        'can_delete_all': can_delete_all,
        'latest_dean_list': latest_dean_list,
    }
    
    return render(request, 'frontend/parking_management.html', context)


@login_required
@require_POST
def delete_parking_application(request, application_id):
    """
    Delete a specific parking application - only for superstaff and specific roles
    """
    from .models import ParkingApplication
    
    # Check if user has permission to delete applications
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions to delete applications'})
    
    try:
        application = ParkingApplication.objects.get(id=application_id)
        application.delete()
        return JsonResponse({'success': True, 'message': 'Parking application deleted successfully'})
    except ParkingApplication.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Parking application not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_all_parking_applications(request):
    """
    Delete all parking applications - only for superstaff and specific roles
    """
    from .models import ParkingApplication
    
    # Check if user has permission
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions'})
    
    try:
        count = ParkingApplication.objects.count()
        ParkingApplication.objects.all().delete()
        return JsonResponse({'success': True, 'message': f'{count} parking applications deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


def exchange_program(request):
    from .models import ExchangeProgramSettings
    
    # Check if exchange program is visible
    settings = ExchangeProgramSettings.get_settings()
    if not settings.is_visible:
        # Show coming soon page
        return render(request, 'frontend/exchange_coming_soon.html', {
            'message': settings.coming_soon_message
        })
    
    nomination_years = ExchangeNomination.YEAR_OF_STUDY_CHOICES
    degree_levels = ExchangeNomination.DEGREE_LEVEL_CHOICES
    semester_choices = ExchangeNomination.SEMESTER_CHOICES
    partner_universities = PartnerUniversity.objects.order_by('name')

    errors = []
    success = False
    default_form_data = {
        'full_name': '',
        'student_email': '',
        'coordinator_name': '',
        'coordinator_email': '',
        'institution': '',
        'year_of_study': '',
        'degree_level': '',
        'semester_to_apply_for': '',
        'academic_year': '',
        'completed_credits': '',
        'total_required_credits': '',
        'major': '',
        'completed_semesters': '',
    }
    form_data = default_form_data.copy()

    if request.method == 'POST':
        form_data = request.POST.dict()
        form_data.pop('csrfmiddlewaretoken', None)

        full_name = form_data.get('full_name', '').strip()
        student_email = form_data.get('student_email', '').strip()
        coordinator_name = form_data.get('coordinator_name', '').strip()
        coordinator_email = form_data.get('coordinator_email', '').strip()
        institution = form_data.get('institution', '').strip()
        year_of_study = form_data.get('year_of_study', '').strip()
        degree_level = form_data.get('degree_level', '').strip()
        semester_to_apply_for = form_data.get('semester_to_apply_for', '').strip()
        academic_year = form_data.get('academic_year', '').strip()
        completed_credits_raw = form_data.get('completed_credits', '').strip()
        total_required_credits_raw = form_data.get('total_required_credits', '').strip()
        major = form_data.get('major', '').strip()
        completed_semesters_raw = form_data.get('completed_semesters', '').strip()

        if not full_name:
            errors.append('Full name is required.')
        if not student_email:
            errors.append('Student email address is required.')
        if not coordinator_name:
            errors.append('Exchange coordinator name is required.')
        if not coordinator_email:
            errors.append('Exchange coordinator email is required.')
        if not institution:
            errors.append('Home institution is required.')
        elif not PartnerUniversity.objects.filter(name=institution).exists():
            errors.append('Selected home institution is not recognized. Please pick from the list of partner universities.')
        if year_of_study not in dict(nomination_years):
            errors.append('Please select a valid year of study.')
        if degree_level not in dict(degree_levels):
            errors.append('Please specify whether the nominee is a bachelor\'s or master\'s student.')
        if semester_to_apply_for not in dict(semester_choices):
            errors.append('Please select the semester the nominee plans to attend.')
        if not academic_year:
            errors.append('The academic year is required (e.g., 2025/2026).')
        else:
            if not re.match(r"^\d{4}/\d{4}$", academic_year):
                errors.append('Academic year must be in the format YYYY/YYYY (e.g., 2025/2026).')
        if not major:
            errors.append('Major in the home institution is required.')

        completed_credits = None
        total_required_credits = None
        completed_semesters = None

        if not completed_credits_raw:
            errors.append('Completed credits are required.')
        else:
            try:
                completed_credits = int(completed_credits_raw)
                if completed_credits < 0:
                    errors.append('Completed credits cannot be negative.')
            except ValueError:
                errors.append('Completed credits must be a number.')

        if not total_required_credits_raw:
            errors.append('Total required credits to graduate are required.')
        else:
            try:
                total_required_credits = int(total_required_credits_raw)
                if total_required_credits <= 0:
                    errors.append('Total required credits must be greater than zero.')
            except ValueError:
                errors.append('Total required credits must be a number.')

        if not completed_semesters_raw:
            errors.append('Number of completed ordinary semesters is required.')
        else:
            try:
                completed_semesters = int(completed_semesters_raw)
                if completed_semesters < 0:
                    errors.append('Completed semesters cannot be negative.')
            except ValueError:
                errors.append('Completed semesters must be a number.')

        if not errors:
            ExchangeNomination.objects.create(
                full_name=full_name,
                student_email=student_email,
                coordinator_name=coordinator_name,
                coordinator_email=coordinator_email,
                institution=institution,
                year_of_study=year_of_study,
                degree_level=degree_level,
                semester_to_apply_for=semester_to_apply_for,
                academic_year=academic_year,
                completed_credits=completed_credits,
                total_required_credits=total_required_credits,
                major=major,
                completed_semesters=completed_semesters,
            )
            success = True
            form_data = default_form_data.copy()

    context = {
        'nomination_years': nomination_years,
        'degree_levels': degree_levels,
        'semester_choices': semester_choices,
        'partner_universities': partner_universities,
        'errors': errors,
        'success': success,
        'form_data': form_data,
    }
    return render(request, 'frontend/exchange_apply.html', context)


def exchange_application(request):
    submitted = False
    if request.method == 'POST':
        form = ExchangeApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            submitted = True
            form = ExchangeApplicationForm()
    else:
        form = ExchangeApplicationForm()

    context = {
        'form': form,
        'submitted': submitted,
        'partner_count': PartnerUniversity.objects.count(),
    }
    return render(request, 'frontend/exchange_application.html', context)


def user_is_exchange_officer(user):
    return user.is_authenticated and (user.is_superuser or getattr(user, 'is_exchange_officer', False))


@login_required
def exchange_dashboard(request):
    if not user_is_exchange_officer(request.user):
        messages.error(request, 'You do not have permission to access the exchange dashboard.')
        return redirect('portal')

    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })

    partner_form = PartnerUniversityForm()
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'add_partner':
            partner_form = PartnerUniversityForm(request.POST, request.FILES)
            if partner_form.is_valid():
                partner_form.save()
                messages.success(request, 'Partner university added successfully.')
                return redirect('exchange_dashboard')
            messages.error(request, 'Unable to add partner university. Please correct the errors below.')

        elif form_type == 'delete_partner':
            partner_id = request.POST.get('partner_id')
            partner = get_object_or_404(PartnerUniversity, pk=partner_id)
            if partner.logo:
                partner.logo.delete(save=False)
            partner.delete()
            messages.success(request, 'Partner university deleted successfully.')
            return redirect('exchange_dashboard')

    active_nominations_qs = ExchangeNomination.objects.filter(is_archived=False).order_by('-submitted_at')
    total_nominations = ExchangeNomination.objects.count()
    active_nominations_count = active_nominations_qs.count()
    archived_nominations_count = ExchangeNomination.objects.filter(is_archived=True).count()

    completion_rates = []
    for nomination in active_nominations_qs:
        if nomination.total_required_credits:
            ratio = nomination.completed_credits / nomination.total_required_credits
            completion_rates.append(ratio * 100)

    average_completion_rate = round(sum(completion_rates) / len(completion_rates), 1) if completion_rates else 0

    semester_choices = dict(ExchangeNomination.SEMESTER_CHOICES)
    degree_choices = dict(ExchangeNomination.DEGREE_LEVEL_CHOICES)

    semester_breakdown = [
        {
            'code': row['semester_to_apply_for'],
            'label': semester_choices.get(row['semester_to_apply_for'], row['semester_to_apply_for']),
            'count': row['total'],
        }
        for row in active_nominations_qs.values('semester_to_apply_for').annotate(total=Count('id')).order_by('semester_to_apply_for')
    ]

    degree_breakdown = [
        {
            'code': row['degree_level'],
            'label': degree_choices.get(row['degree_level'], row['degree_level']),
            'count': row['total'],
        }
        for row in active_nominations_qs.values('degree_level').annotate(total=Count('id')).order_by('degree_level')
    ]

    partner_count = PartnerUniversity.objects.count()
    recent_partners = PartnerUniversity.objects.order_by('-created_at')[:6]
    partner_directory = PartnerUniversity.objects.order_by('name')

    application_count = ExchangeApplication.objects.count()
    recent_applications = ExchangeApplication.objects.order_by('-submitted_at')[:5]

    context = {
        'stats': {
            'total_nominations': total_nominations,
            'active_nominations': active_nominations_count,
            'archived_nominations': archived_nominations_count,
            'partner_count': partner_count,
            'application_count': application_count,
            'average_completion_rate': average_completion_rate,
        },
        'recent_nominations': list(active_nominations_qs[:5]),
        'recent_partners': recent_partners,
        'partner_directory': partner_directory,
        'recent_applications': list(recent_applications),
        'semester_breakdown': semester_breakdown,
        'degree_breakdown': degree_breakdown,
        'partner_form': partner_form,
    }

    return render(request, 'frontend/exchange_dashboard.html', context)


@login_required
def exchange_officer_dashboard(request):
    if not user_is_exchange_officer(request.user):
        return HttpResponseForbidden("You do not have permission to view this page.")

    show_archived = request.GET.get('show') == 'archived'
    nominations = ExchangeNomination.objects.filter(is_archived=show_archived).order_by('-submitted_at')
    stats = {
        'active': ExchangeNomination.objects.filter(is_archived=False).count(),
        'archived': ExchangeNomination.objects.filter(is_archived=True).count(),
    }

    context = {
        'nominations': nominations,
        'show_archived': show_archived,
        'stats': stats,
    }
    return render(request, 'frontend/exchange_officer_dashboard.html', context)


@login_required
def exchange_application_management(request):
    if not user_is_exchange_officer(request.user):
        return HttpResponseForbidden("You do not have permission to view this page.")

    show_archived = request.GET.get('show') == 'archived'
    applications = ExchangeApplication.objects.filter(is_archived=show_archived).select_related('home_institution').order_by('-submitted_at')
    stats = {
        'active': ExchangeApplication.objects.filter(is_archived=False).count(),
        'archived': ExchangeApplication.objects.filter(is_archived=True).count(),
    }

    context = {
        'applications': applications,
        'show_archived': show_archived,
        'stats': stats,
    }
    return render(request, 'frontend/exchange_application_management.html', context)


@login_required
@require_POST
def toggle_exchange_nomination_archive(request, nomination_id):
    if not user_is_exchange_officer(request.user):
        return HttpResponseForbidden("You do not have permission to perform this action.")

    nomination = get_object_or_404(ExchangeNomination, pk=nomination_id)
    action = request.POST.get('action')

    if action == 'archive':
        nomination.is_archived = True
        message_text = 'Nomination archived successfully.'
    else:
        nomination.is_archived = False
        message_text = 'Nomination restored successfully.'

    nomination.save(update_fields=['is_archived'])
    messages.success(request, message_text)

    next_url = request.POST.get('next') or reverse('exchange_officer_dashboard')
    return redirect(next_url)


@login_required
@require_POST
def toggle_exchange_application_archive(request, application_id):
    if not user_is_exchange_officer(request.user):
        return HttpResponseForbidden("You do not have permission to perform this action.")

    application = get_object_or_404(ExchangeApplication, pk=application_id)
    action = request.POST.get('action')

    if action == 'archive':
        application.is_archived = True
        message_text = 'Application archived successfully.'
    else:
        application.is_archived = False
        message_text = 'Application restored successfully.'

    application.save(update_fields=['is_archived'])
    messages.success(request, message_text)

    next_url = request.POST.get('next') or reverse('exchange_application_management')
    return redirect(next_url)


def members(request):
    members = User.objects.filter(is_member=True)
    # order members by role, president first, vice president second, etc.
    new_order = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY', 'TREASURER', 'MEMBER']
    members = sorted(members, key=lambda x: new_order.index(x.role) if x.role in new_order else 100)
    return render(request, 'frontend/members.html', {'members': members})

@login_required
def portal(request):
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        # Render a "desktop only" message page
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # user_count is now automatically available via context processor
    # include events for portal users
    try:
        from .models import Event
        events = Event.objects.all().order_by('-start_date')
    except Exception:
        events = []

    return render(request, 'frontend/portal.html', {'user': request.user, 'events': events})


@login_required
def refresh_courses(request):
    """Admin view to refresh course data from CMU website"""
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    
    # Check if user has permission (only president and staff can access)
    if not (request.user.role == 'PRESIDENT' or request.user.is_staff):
        return HttpResponseForbidden("Access denied. President or staff access required.")
    
    if request.method == 'POST':
        try:
            from django.core.management import call_command
            from .models import Course
            import io
            from contextlib import redirect_stdout
            
            # Capture the output of the management command
            output = io.StringIO()
            
            # Delete all existing courses
            deleted_count = Course.objects.count()
            Course.objects.all().delete()
            
            # Run the fetch_courses command
            with redirect_stdout(output):
                call_command('fetch_courses')
            
            command_output = output.getvalue()
            
            # Get the new course count
            new_count = Course.objects.count()
            
            # Run additional cleanup to ensure no invalid course IDs remain
            cleanup_count = cleanup_invalid_course_ids()
            final_count = Course.objects.count()
            
            success_message = f"Course refresh completed successfully!\nDeleted: {deleted_count} old courses\nFetched: {new_count} new courses"
            
            if cleanup_count > 0:
                success_message += f"\nCleaned up: {cleanup_count} invalid course IDs\nFinal count: {final_count} valid courses"
            
            messages.success(request, success_message)
            
            # Log the refresh action
            print(f"Course refresh by {request.user.username}: {deleted_count} deleted, {new_count} fetched, {cleanup_count} cleaned up")
            
        except Exception as e:
            messages.error(
                request,
                f"Error refreshing courses: {str(e)}"
            )
    
    # Get current course statistics
    try:
        from .models import Course
        total_courses = Course.objects.count()
        # Get unique departments properly using set() to avoid duplicates
        departments_queryset = Course.objects.values_list('department', flat=True)
        unique_departments = sorted(set([dept for dept in departments_queryset if dept]))
        active_courses = Course.objects.filter(is_active=True).count()
        courses_with_syllabus = Course.objects.filter(syllabus_url__isnull=False).count()
    except Exception as e:
        # Handle case where Course model doesn't exist or database table hasn't been created
        total_courses = 0
        unique_departments = []
        active_courses = 0
        courses_with_syllabus = 0
    
    context = {
        'total_courses': total_courses,
        'departments': unique_departments,
        'active_courses': active_courses,
        'courses_with_syllabus': courses_with_syllabus,
        'user': request.user,
    }
    
    return render(request, 'frontend/refresh_courses.html', context)


def login_view(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("portal"))
        else:
            return render(request, "frontend/login.html", {
                "message": "Invalid username and/or password."
            })
    else:
        if request.user.is_authenticated:
            return HttpResponseRedirect(reverse("portal"))
        return render(request, "frontend/login.html")


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("home"))



@login_required
def add_member(request):
    """
    Add new member functionality - accessible to all users but only functional for authorized roles
    Supports both single member and bulk member creation
    """
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # Check if user has permission to add members
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'MANAGER']
    can_add_member = request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles
    
    message = None
    error_message = None
    bulk_results = None
    
    if request.method == "POST" and can_add_member:
        # Check if this is a bulk creation request
        bulk_data = request.POST.get("bulk_data", "").strip()
        
        if bulk_data:
            # Handle bulk member creation
            bulk_results = process_bulk_member_creation(bulk_data)
            if bulk_results['success_count'] > 0:
                message = f"Successfully created {bulk_results['success_count']} member(s)."
            if bulk_results['error_count'] > 0:
                error_message = f"Failed to create {bulk_results['error_count']} member(s). See details below."
        else:
            # Handle single member creation (existing logic)
            username = request.POST.get("username", "").strip()
            password = request.POST.get("password", "").strip()
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            role = request.POST.get("role", "MEMBER")
            is_member = request.POST.get("is_member") == "on"
            
            # Validation
            if not all([username, password, first_name, last_name]):
                error_message = "All fields are required."
            elif len(password) < 6:
                error_message = "Password must be at least 6 characters long."
            elif User.objects.filter(username=username).exists():
                error_message = "Username already exists. Please choose a different username."
            else:
                try:
                    # Create new user with hardcoded email
                    user = User.objects.create_user(
                        username=username,
                        email="sss@sss.com",
                        password=password,
                        first_name=first_name,
                        last_name=last_name
                    )
                    user.role = role
                    user.is_member = is_member
                    user.save()
                    
                    message = f"Member '{first_name} {last_name}' has been successfully created with username '{username}'."
                    
                except IntegrityError:
                    error_message = "An error occurred while creating the user. Username might already exist."
                except Exception as e:
                    error_message = f"An unexpected error occurred: {str(e)}"
    
    elif request.method == "POST" and not can_add_member:
        error_message = "You don't have permission to add new members."
    
    # Get role choices for the dropdown
    role_choices = [
        ('PRESIDENT', 'President'),
        ('VICE_PRESIDENT', 'Vice President'),
        ('SECRETARY', 'Secretary'),
        ('TREASURER', 'Treasurer'),
        ('MANAGER', 'Manager'),
        ('MEMBER', 'Member'),
    ]
    
    context = {
        'can_add_member': can_add_member,
        'message': message,
        'error_message': error_message,
        'bulk_results': bulk_results,
        'role_choices': role_choices,
        'user': request.user,
    }
    
    return render(request, 'frontend/add_member.html', context)


def process_bulk_member_creation(bulk_data):
    """
    Process bulk member creation from multiline string format:
    username,password,First_name Last_name
    
    Example:
    alaadlc,mar7badawli,Alaa El Haj
    mai.Abd,mis2025,Maida Abdullah
    
    Returns dict with success_count, error_count, successes list, and errors list
    """
    results = {
        'success_count': 0,
        'error_count': 0,
        'successes': [],
        'errors': []
    }
    
    lines = bulk_data.strip().split('\n')
    
    for line_num, line in enumerate(lines, start=1):
        line = line.strip()
        if not line:
            continue  # Skip empty lines
        
        parts = [p.strip() for p in line.split(',')]
        
        if len(parts) != 3:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': f'Invalid format. Expected 3 comma-separated values, got {len(parts)}.'
            })
            results['error_count'] += 1
            continue
        
        username, password, full_name = parts
        
        # Split full name into first and last name
        name_parts = full_name.split()
        if len(name_parts) < 2:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': 'Full name must contain at least first and last name.'
            })
            results['error_count'] += 1
            continue
        
        first_name = name_parts[0]
        last_name = ' '.join(name_parts[1:])  # Everything after first name is last name
        
        # Validate
        if not username or not password or not first_name or not last_name:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': 'Username, password, and full name cannot be empty.'
            })
            results['error_count'] += 1
            continue
        
        if len(password) < 6:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': 'Password must be at least 6 characters long.'
            })
            results['error_count'] += 1
            continue
        
        if User.objects.filter(username=username).exists():
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': f'Username "{username}" already exists.'
            })
            results['error_count'] += 1
            continue
        
        # Create the user
        try:
            user = User.objects.create_user(
                username=username,
                email="sss@sss.com",
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            user.role = 'MEMBER'
            user.is_member = True
            user.save()
            
            results['successes'].append({
                'line': line_num,
                'username': username,
                'name': f'{first_name} {last_name}'
            })
            results['success_count'] += 1
            
        except IntegrityError:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': f'Database error: Username "{username}" might already exist.'
            })
            results['error_count'] += 1
        except Exception as e:
            results['errors'].append({
                'line': line_num,
                'data': line,
                'error': f'Unexpected error: {str(e)}'
            })
            results['error_count'] += 1
    
    return results


@login_required
def manage_members(request):
    """
    Manage members functionality - accessible only to president and staff
    Allows deleting members and changing their roles
    """
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # Check if user has permission to manage members
    allowed_roles = ['PRESIDENT']
    can_manage = request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles
    
    if not can_manage:
        return render(request, 'frontend/access_denied.html', {
            'message': 'Only President and Staff can manage members.',
            'required_roles': ['President', 'Staff']
        })
    
    # Get all users ordered by role hierarchy
    role_order = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY', 'TREASURER', 'MANAGER', 'MEMBER']
    all_users = User.objects.all().order_by('is_staff', 'is_superuser')
    
    # Sort users by role hierarchy
    users = sorted(all_users, key=lambda x: (
        0 if x.is_superuser else 1,
        0 if x.is_staff else 1,
        role_order.index(x.role) if x.role in role_order else 100,
        x.first_name.lower() if x.first_name else x.username.lower()
    ))
    
    # Get role choices for the dropdown
    role_choices = [
        ('PRESIDENT', 'President'),
        ('VICE_PRESIDENT', 'Vice President'),
        ('SECRETARY', 'Secretary'),
        ('TREASURER', 'Treasurer'),
        ('MANAGER', 'Manager'),
        ('MEMBER', 'Member'),
    ]
    
    # Calculate statistics
    total_users = len(users)
    active_members = len([u for u in users if u.is_member])
    staff_users = len([u for u in users if u.is_staff or u.is_superuser])
    
    context = {
        'users': users,
        'role_choices': role_choices,
        'total_users': total_users,
        'active_members': active_members,
        'staff_users': staff_users,
        'current_user': request.user,
    }
    
    return render(request, 'frontend/manage_members.html', context)


@login_required
@require_POST
def update_member_role(request, user_id):
    """
    Update a member's role - only for president and staff
    """
    # Check if user has permission
    allowed_roles = ['PRESIDENT']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions'})
    
    try:
        member = User.objects.get(id=user_id)
        new_role = request.POST.get('role')
        new_is_member = request.POST.get('is_member') == 'true'
        
        # Prevent users from modifying superusers unless they are superuser themselves
        if member.is_superuser and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Cannot modify superuser accounts'})
        
        # Prevent users from removing the last president (unless they are superuser)
        if member.role == 'PRESIDENT' and new_role != 'PRESIDENT' and not request.user.is_superuser:
            president_count = User.objects.filter(role='PRESIDENT').count()
            if president_count <= 1:
                return JsonResponse({'success': False, 'message': 'Cannot remove the last president'})
        
        # Update the member
        member.role = new_role
        member.is_member = new_is_member
        member.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully updated {member.first_name} {member.last_name}\'s role to {new_role}'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'User not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_member(request, user_id):
    """
    Delete a member - only for president and staff
    """
    # Check if user has permission
    allowed_roles = ['PRESIDENT']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions'})
    
    try:
        member = User.objects.get(id=user_id)
        
        # Prevent deletion of self
        if member.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'Cannot delete your own account'})
        
        # Prevent deletion of superusers unless the current user is superuser
        if member.is_superuser and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Cannot delete superuser accounts'})
        
        # Prevent deletion of the last president (unless current user is superuser)
        if member.role == 'PRESIDENT' and not request.user.is_superuser:
            president_count = User.objects.filter(role='PRESIDENT').count()
            if president_count <= 1:
                return JsonResponse({'success': False, 'message': 'Cannot delete the last president'})
        
        member_name = f"{member.first_name} {member.last_name}"
        member.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully deleted {member_name}'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'User not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@user_passes_test(lambda u: u.is_staff)
def register(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = "sss@sss.com"
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]


        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "frontend/register.html", {
                "message": "Passwords must match."
            })

        try:
            user = User.objects.create_user(username, email, password, first_name=first_name, last_name=last_name)
            user.save()
        except IntegrityError:
            return render(request, "frontend/register.html", {
                "message": "Username is already previously taken."
            })
        return HttpResponseRedirect(reverse("home"))
    else:
        if request.user.is_staff:
            return render(request, "frontend/register.html")



@login_required
def newdl(request):
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        # Render a "desktop only" message page
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    if request.method == 'POST':
        semester = request.POST.get('semester')
        year = request.POST.get('year')
        excel_file = request.FILES.get('excel_file')   
        if not semester or not year or not excel_file:
            context = {
                'message': 'All fields are required.',
                'years': range(2007, 2046) 
            }
            return render(request, 'frontend/newdl.html', context)
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            context = {
                'message': 'Please upload a valid Excel file (.xlsx or .xls).',
                'years': range(2007, 2046)
            }
            return render(request, 'frontend/newdl.html', context)      
        try:
            file_extension = os.path.splitext(excel_file.name)[1]  
            new_filename = f"{semester}{year}dean list{file_extension}"        
            file_content = excel_file.read()
            renamed_file = ContentFile(file_content, name=new_filename)
            dean_list = DeanList.objects.create(
                semester=semester,
                year=int(year),
                excel_file=renamed_file
            )
            dean_list.save()
            excel_file.seek(0)
            df, header_row, students_saved = process_dean_list_excel(excel_file, semester, year)
            print(f"Processing completed: header at row {header_row}, {students_saved} students saved")

            return render(request, 'frontend/newdl.html', {
                'message': f'Dean\'s list created successfully. {students_saved} students imported.',
                'years': range(2007, 2046)
            }) 
            
        except Exception as e:
            context = {
                'message': f'Error creating dean\'s list: {str(e)}',
                'years': range(2007, 2046)
            }
            return render(request, 'frontend/newdl.html', context)

    context = {
        'years': range(2007, 2046)
    }
    return render(request, 'frontend/newdl.html', context)


def deanslist(request):
    """
    Display dean's list students filtered by semester and year.
    Group by major and order by GPA (descending) within each major.
    """
    # Get available semesters and years from the database
    available_lists = DeanList.objects.values('semester', 'year').distinct().order_by('-year', 'semester')
    
    # Initialize variables
    students_by_major = {}
    selected_semester = request.GET.get('semester', '')
    selected_year = request.GET.get('year', '')
    selected_major = request.GET.get('major', '')
    available_majors = []
    
    # If semester and year are provided, get the students
    if selected_semester and selected_year:
        try:
            year_int = int(selected_year)
            students = DeanListStudent.objects.filter(
                semester=selected_semester,
                year=year_int
            ).order_by('student_major', '-gpa', '-passed_credits', 'student_name')
            
            # Get unique majors for this specific semester/year
            available_majors = students.values_list('student_major', flat=True).distinct().order_by('student_major')
            available_majors = [major for major in available_majors if major and major.strip()]  # Remove empty majors
            
            # Group students by major
            for student in students:
                major = student.student_major
                if not major or not major.strip():
                    major = 'Unknown'
                else:
                    major = major.strip()
                    
                if major not in students_by_major:
                    students_by_major[major] = []
                students_by_major[major].append(student)
            
            # If a specific major is selected, filter to show only that major
            if selected_major and selected_major.strip():
                print(f"DEBUG: Selected major: '{selected_major}'")
                print(f"DEBUG: Available majors in students_by_major: {list(students_by_major.keys())}")
                if selected_major in students_by_major:
                    students_by_major = {selected_major: students_by_major[selected_major]}
                    print(f"DEBUG: Filtered to show only {selected_major}: {len(students_by_major[selected_major])} students")
                else:
                    # Selected major not found for this semester/year
                    students_by_major = {}
                    print(f"DEBUG: Selected major '{selected_major}' not found for {selected_semester} {selected_year}")
                    
        except ValueError:
            # Invalid year format
            students_by_major = {}
            available_majors = []
    else:
        # Get all unique majors from all records for the dropdown when no semester/year selected
        all_majors = DeanListStudent.objects.values_list('student_major', flat=True).distinct().order_by('student_major')
        available_majors = [major for major in all_majors if major and major.strip()]
    
    # Get available years from the database
    available_years = DeanListStudent.objects.values_list('year', flat=True).distinct().order_by('year')
    
    context = {
        'available_lists': available_lists,
        'all_majors': sorted(available_majors),
        'students_by_major': students_by_major,
        'selected_semester': selected_semester,
        'selected_year': selected_year,
        'selected_major': selected_major,
        'years': sorted(available_years),
        'total_students': sum(len(students) for students in students_by_major.values()),
    }
    
    return render(request, 'frontend/deanslist.html', context)


@login_required
def student_search(request):
    """
    Search for students by ID or name and display their dean's list rankings
    """
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    search_results = []
    search_query = ""
    search_type = ""
    multiple_matches = False
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id', '').strip()
        student_name = request.POST.get('student_name', '').strip()
        
        if student_id:
            # Search by student ID (exact match)
            search_query = student_id
            search_type = "ID"
            students = DeanListStudent.objects.filter(student_id=student_id).order_by('year', 'semester')
            
            if students.exists():
                search_results = calculate_student_rankings(students, request.user)
            
        elif student_name:
            # Search by name (substring matching with better Arabic support)
            search_query = student_name
            search_type = "name"
            
            # Split the search query into individual words for better matching
            search_words = [word.strip() for word in student_name.split() if word.strip()]
            
            if search_words:
                # Build a query that matches all words in any order within the name
                query = Q()
                for word in search_words:
                    query &= Q(student_name__icontains=word)
                
                students = DeanListStudent.objects.filter(query).order_by('student_name', 'year', 'semester')
            else:
                students = DeanListStudent.objects.none()
            
            if students.exists():
                # Group by student name to handle multiple matches
                students_by_name = {}
                for student in students:
                    name = student.student_name
                    if name not in students_by_name:
                        students_by_name[name] = []
                    students_by_name[name].append(student)
                
                # If multiple different names match, show selection
                if len(students_by_name) > 1:
                    multiple_matches = True
                    search_results = [
                        {
                            'student_name': name,
                            'student_id': students_list[0].student_id,
                            'appearances': len(students_list),
                            'years': sorted(list(set(f"{s.semester} {s.year}" for s in students_list)))
                        }
                        for name, students_list in students_by_name.items()
                    ]
                else:
                    # Single name match, show full rankings
                    all_students = list(students)
                    search_results = calculate_student_rankings(all_students, request.user)
    
    context = {
        'search_results': search_results,
        'search_query': search_query,
        'search_type': search_type,
        'multiple_matches': multiple_matches,
    }
    
    return render(request, 'frontend/student_search.html', context)


def calculate_student_rankings(students, user):
    """
    Calculate rankings for a student across different semesters/years
    """
    results = []
    
    # Group by semester and year
    by_semester_year = {}
    for student in students:
        key = f"{student.semester}_{student.year}"
        if key not in by_semester_year:
            by_semester_year[key] = []
        by_semester_year[key].append(student)
    
    # For each semester/year, calculate the student's ranking
    for semester_year_key, student_list in by_semester_year.items():
        student = student_list[0]  # Should be only one per semester/year
        semester = student.semester
        year = student.year
        
        # Get all students for this semester/year, ordered by GPA descending
        all_students_in_semester = DeanListStudent.objects.filter(
            semester=semester,
            year=year
        ).order_by('-gpa', 'student_name')
        
        # Find the overall ranking (1-based)
        overall_ranking = 1
        for idx, s in enumerate(all_students_in_semester):
            if s.student_id == student.student_id:
                overall_ranking = idx + 1
                break
        
        total_students = all_students_in_semester.count()
        
        # Calculate ranking within major
        students_in_major = DeanListStudent.objects.filter(
            semester=semester,
            year=year,
            student_major=student.student_major
        ).order_by('-gpa', 'student_name')
        
        # Find the major ranking (1-based)
        major_ranking = 1
        for idx, s in enumerate(students_in_major):
            if s.student_id == student.student_id:
                major_ranking = idx + 1
                break
        
        total_students_in_major = students_in_major.count()
        
        # Only show GPA if user is superuser/staff
        show_gpa = user.is_superuser or user.is_staff
        
        results.append({
            'student': student,
            'ranking': overall_ranking,
            'total_students': total_students,
            'major_ranking': major_ranking,
            'total_students_in_major': total_students_in_major,
            'semester_display': f"{semester.title()} {year}",
            'percentage_rank': round((overall_ranking / total_students) * 100, 1) if total_students > 0 else 0,
            'major_percentage_rank': round((major_ranking / total_students_in_major) * 100, 1) if total_students_in_major > 0 else 0,
            'show_gpa': show_gpa
        })
    
    # Sort by year and semester (most recent first)
    results.sort(key=lambda x: (x['student'].year, x['student'].semester), reverse=True)
    
    return results


@login_required
def application_management(request):
    """
    Manage membership applications - view all applications with automatic rejection flags
    """
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # Get all applications ordered by eligibility first, then by GPA and passed credits
    all_applications = Application.objects.all()
    
    # Separate eligible and rejected applications
    eligible_apps = []
    rejected_apps = []
    
    for app in all_applications:
        auto_reject = app.GPA < Decimal('3.5') or app.passed_credits < 30
        if auto_reject:
            rejected_apps.append(app)
        else:
            eligible_apps.append(app)
    
    # Sort eligible applications by GPA (highest first), then by passed credits (highest first)
    eligible_apps.sort(key=lambda x: (-x.GPA, -x.passed_credits, x.time_submitted))
    
    # Sort rejected applications by GPA (highest first), then by passed credits (highest first)
    rejected_apps.sort(key=lambda x: (-x.GPA, -x.passed_credits, x.time_submitted))
    
    # Combine lists with eligible first, then rejected
    applications = eligible_apps + rejected_apps
    
    # Add rejection flags and eligibility status to each application
    applications_with_status = []
    separator_added = False
    
    for app in applications:
        # Determine if application should be automatically rejected
        auto_reject = app.GPA < Decimal('3.5') or app.passed_credits < 30
        
        # Add separator before first rejected application
        add_separator = auto_reject and not separator_added and len([a for a in applications if not (a.GPA < Decimal('3.5') or a.passed_credits < 30)]) > 0
        if add_separator:
            separator_added = True
        
        # Determine rejection reasons
        rejection_reasons = []
        if app.GPA < Decimal('3.5'):
            rejection_reasons.append(f"GPA below 3.5 ({app.GPA})")
        if app.passed_credits < 30:
            rejection_reasons.append(f"Insufficient credits ({app.passed_credits} < 30)")
        
        # Get major display name
        major_display = dict(majors).get(app.major, app.major)
        
        applications_with_status.append({
            'application': app,
            'auto_reject': auto_reject,
            'rejection_reasons': rejection_reasons,
            'major_display': major_display,
            'eligible': not auto_reject,
            'add_separator': add_separator
        })
    
    # Calculate statistics
    total_applications = len(applications_with_status)
    eligible_applications = len([app for app in applications_with_status if app['eligible']])
    rejected_applications = total_applications - eligible_applications
    
    # Check if user can delete applications (individual and all)
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    can_delete = request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles
    can_delete_all = can_delete  # Same permissions for both
    
    context = {
        'applications_with_status': applications_with_status,
        'total_applications': total_applications,
        'eligible_applications': eligible_applications,
        'rejected_applications': rejected_applications,
        'can_delete': can_delete,
        'can_delete_all': can_delete_all,
    }
    
    return render(request, 'frontend/application_management.html', context)


@login_required
@require_POST
def delete_application(request, application_id):
    """
    Delete a specific application - only for superstaff and specific roles
    """
    # Check if user has permission to delete applications
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions to delete applications'})
    
    try:
        application = Application.objects.get(id=application_id)
        application.delete()
        return JsonResponse({'success': True, 'message': 'Application deleted successfully'})
    except Application.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Application not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_all_applications(request):
    """
    Delete all applications - only for superstaff and specific roles
    """
    # Check if user has permission
    allowed_roles = ['PRESIDENT', 'VICE_PRESIDENT', 'SECRETARY']
    if not (request.user.is_superuser or request.user.is_staff or request.user.role in allowed_roles):
        return JsonResponse({'success': False, 'message': 'Insufficient permissions'})
    
    try:
        count = Application.objects.count()
        Application.objects.all().delete()
        return JsonResponse({'success': True, 'message': f'{count} applications deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def thread_settings(request):
    """
    Manage thread settings - accessible to any logged-in user
    """
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    # Any logged-in user can manage thread settings
    settings = ThreadSettings.get_settings()
    message = None
    error_message = None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_threads':
            settings.allow_new_threads = not settings.allow_new_threads
            settings.updated_by = request.user
            settings.save()
            
            status = "enabled" if settings.allow_new_threads else "disabled"
            message = f"Thread creation has been {status}."
            
        elif action == 'update_message':
            new_message = request.POST.get('closure_message', '').strip()
            if new_message:
                settings.closure_message = new_message
                settings.updated_by = request.user
                settings.save()
                message = "Closure message updated successfully."
            else:
                error_message = "Closure message cannot be empty."
    
    # Get thread statistics
    total_threads = Thread.objects.count()
    total_replies = Reply.objects.count()
    recent_threads = Thread.objects.filter(
        created_at__gte=datetime.now() - timedelta(days=7)
    ).count()
    
    context = {
        'settings': settings,
        'message': message,
        'error_message': error_message,
        'total_threads': total_threads,
        'total_replies': total_replies,
        'recent_threads': recent_threads,
    }
    
    return render(request, 'frontend/thread_settings.html', context)


@login_required
def exchange_program_settings(request):
    """
    Manage exchange program visibility settings - accessible to logged-in users
    """
    from .models import ExchangeProgramSettings
    
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    settings = ExchangeProgramSettings.get_settings()
    message = None
    error_message = None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_visibility':
            settings.is_visible = not settings.is_visible
            settings.updated_by = request.user
            settings.save()
            
            status = "visible" if settings.is_visible else "hidden"
            message = f"Exchange program has been {status}."
            
        elif action == 'update_message':
            new_message = request.POST.get('coming_soon_message', '').strip()
            if new_message:
                settings.coming_soon_message = new_message
                settings.updated_by = request.user
                settings.save()
                message = "Coming soon message updated successfully."
            else:
                error_message = "Coming soon message cannot be empty."
    
    context = {
        'settings': settings,
        'message': message,
        'error_message': error_message,
    }
    
    return render(request, 'frontend/exchange_program_settings.html', context)


def contact(request):
    """
    Contact page with thread system - accessible to everyone (logged in or not)
    Shows all threads and allows creating new threads (if enabled)
    """
    settings = ThreadSettings.get_settings()
    
    if request.method == 'POST':
        # Check if thread creation is allowed
        if not settings.allow_new_threads:
            context = {
                'error_message': settings.closure_message,
                'title': request.POST.get('title', ''),
                'content': request.POST.get('content', ''),
                'threads_disabled': True,
                'settings': settings
            }
            return render(request, 'frontend/contact.html', context)
        
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        
        if title and content:
            # Create new thread
            thread = Thread.objects.create(
                title=title,
                content=content
            )
            return redirect('contact')
        else:
            # Handle validation error
            context = {
                'error_message': 'Both title and content are required.',
                'title': title,
                'content': content,
                'settings': settings
            }
            return render(request, 'frontend/contact.html', context)
    
    # Get all threads ordered by creation date (newest first)
    all_threads = Thread.objects.all().order_by('-created_at')
    
    # Pagination - 6 threads per page
    paginator = Paginator(all_threads, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add reply count, latest reply info, and first reply to each thread
    threads_with_info = []
    for thread in page_obj:
        reply_count = thread.replies.count()
        latest_reply = thread.get_latest_reply()
        first_reply = thread.replies.order_by('created_at').first()
        
        threads_with_info.append({
            'thread': thread,
            'reply_count': reply_count,
            'latest_reply': latest_reply,
            'first_reply': first_reply,
        })
    
    context = {
        'threads_with_info': threads_with_info,
        'page_obj': page_obj,
        'user_is_staff': request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser),
        'user_can_delete': request.user.is_authenticated and (request.user.is_member or request.user.is_staff or request.user.is_superuser),
        'settings': settings,
    }
    
    return render(request, 'frontend/contact.html', context)


def thread_detail(request, thread_id):
    """
    View individual thread with all replies
    """
    thread = get_object_or_404(Thread, id=thread_id)
    replies = thread.replies.all().order_by('created_at')
    
    context = {
        'thread': thread,
        'replies': replies,
        'user_can_reply': request.user.is_authenticated and isinstance(request.user, User),
        'user_can_delete': request.user.is_authenticated and (request.user.is_member or request.user.is_staff or request.user.is_superuser),
    }
    
    return render(request, 'frontend/thread_detail.html', context)


@login_required
@require_POST
def add_reply(request, thread_id):
    """
    Add a reply to a thread - only for logged in Users
    """
    thread = get_object_or_404(Thread, id=thread_id)
    content = request.POST.get('content', '').strip()
    
    if not content:
        return JsonResponse({'success': False, 'message': 'Reply content is required'})
    
    # Only allow Users (from our custom model) to reply
    if not isinstance(request.user, User):
        return JsonResponse({'success': False, 'message': 'Only registered members can reply'})
    
    try:
        reply = Reply.objects.create(
            thread=thread,
            user=request.user,
            content=content
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Reply added successfully',
            'reply_id': reply.id,
            'user_name': request.user.get_full_name(),
            'created_at': reply.created_at.strftime('%B %d, %Y at %I:%M %p'),
            'content': reply.content
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_thread(request, thread_id):
    """
    Delete a thread and all its replies - only for members
    """
    # Check if user is a member
    if not (request.user.is_member or request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Only members can delete threads'})
    
    try:
        thread = get_object_or_404(Thread, id=thread_id)
        thread_title = thread.title
        
        # Delete the thread (this will also delete all replies due to cascade)
        thread.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Thread "{thread_title}" and all its replies have been deleted successfully'
        })
        
    except Thread.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Thread not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def delete_reply(request, reply_id):
    """
    Delete a specific reply - only for members
    """
    # Check if user is a member
    if not (request.user.is_member or request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Only members can delete replies'})
    
    try:
        reply = get_object_or_404(Reply, id=reply_id)
        thread_id = reply.thread.id
        
        # Delete the reply
        reply.delete()
        
        return JsonResponse({
            'success': True, 
            'message': 'Reply deleted successfully',
            'thread_id': thread_id
        })
        
    except Reply.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Reply not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


def custom_404_view(request, exception=None):
    """
    Custom 404 error handler - works both as error handler and regular view
    """
    return render(request, 'frontend/404.html', status=404)


def events_list(request):
    """List upcoming and past events"""
    try:
        from .models import Event
        # Show events in the order they were added (newest first)
        qs = Event.objects.all().order_by('-created_at')
        paginator = Paginator(qs, 9)  # 9 events per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        events = page_obj.object_list
    except Exception:
        page_obj = None
        events = []

    return render(request, 'frontend/events_list.html', {'events': events, 'page_obj': page_obj})


def event_detail(request, pk):
    """Show event details"""
    from django.shortcuts import get_object_or_404
    from datetime import datetime
    try:
        from .models import Event
        event = get_object_or_404(Event, pk=pk)
    except Exception:
        event = None

    # Check if registration is allowed (allowed before and during event, blocked after end)
    registration_allowed = False
    if event:
        now = datetime.utcnow()
        end_date = event.end_date or event.start_date
        end_time = event.end_time or event.start_time or datetime.max.time()
        event_end = datetime.combine(end_date, end_time)
        
        # Allow registration before event starts and during event, block after event ends
        registration_allowed = now <= event_end

    # If the attendance form was submitted, handle it here
    if request.method == 'POST':
        # Check if registration is allowed
        if not registration_allowed:
            messages.error(request, 'Registration is closed. This event has already ended.')
            return render(request, 'frontend/event_detail.html', {'event': event, 'registration_allowed': registration_allowed})
        
        try:
            from .models import Attendance
        except Exception:
            messages.error(request, 'Attendance model not available')
            return render(request, 'frontend/event_detail.html', {'event': event, 'registration_allowed': registration_allowed})

        student_id = request.POST.get('student_id', '').strip()
        student_name = request.POST.get('student_name', '').strip()
        student_email = request.POST.get('email', '').strip()

        if not student_id or not student_name:
            messages.error(request, 'Student ID and name are required.')
            return render(request, 'frontend/event_detail.html', {'event': event, 'registration_allowed': registration_allowed})

        # If an attendance record for this event+student exists, return the same code
        existing = Attendance.objects.filter(event=event, student_id=student_id).first()
        if existing:
            code = existing.code
            return render(request, 'frontend/event_code.html', {'event': event, 'student_id': student_id, 'student_name': student_name, 'code': code, 'existing': True})

        # Generate and persist a unique 6-digit code for this event in a DB-safe way.
        # We try to create the Attendance with a candidate code; on IntegrityError
        # we retry (this avoids the check-then-create race condition).
        import random
        from django.db import IntegrityError as DjangoIntegrityError
        from datetime import datetime

        code = None
        max_attempts = 20
        for attempt in range(max_attempts):
            candidate = f"{random.randint(0, 999999):06d}"
            try:
                attendance = Attendance.objects.create(
                    event=event,
                    student_id=student_id,
                    student_name=student_name,
                    email=student_email or None,
                    code=candidate,
                )
                code = candidate
                created = True
                break
            except DjangoIntegrityError:
                # Could be a collision on (event, code) or (event, student_id).
                # If student record appeared concurrently, reuse it.
                existing = Attendance.objects.filter(event=event, student_id=student_id).first()
                if existing:
                    code = existing.code
                    return render(request, 'frontend/event_code.html', {'event': event, 'student_id': student_id, 'student_name': student_name, 'code': code, 'existing': True})
                # Otherwise assume code collision and retry
                continue
        else:
            # fallback to a timestamp-based code if we couldn't get a unique 6-digit code
            candidate = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            try:
                attendance = Attendance.objects.create(
                    event=event,
                    student_id=student_id,
                    student_name=student_name,
                    email=student_email or None,
                    code=candidate,
                )
                code = candidate
            except Exception:
                # If still failing, try to find an existing record (last resort)
                existing = Attendance.objects.filter(event=event, student_id=student_id).first()
                if existing:
                    code = existing.code
                    return render(request, 'frontend/event_code.html', {'event': event, 'student_id': student_id, 'student_name': student_name, 'code': code, 'existing': True})
                messages.error(request, 'Could not register attendance at this time.')
                return render(request, 'frontend/event_detail.html', {'event': event, 'registration_allowed': registration_allowed})

        return render(request, 'frontend/event_code.html', {'event': event, 'student_id': student_id, 'student_name': student_name, 'code': code, 'existing': False})

    return render(request, 'frontend/event_detail.html', {'event': event, 'registration_allowed': registration_allowed})


def verify_start(request, a, b, c):
    """Dynamic route: given three preset numbers, find the event with those secrets,
    render a template to enter the attendee's code, and on POST set present_start=True
    if (and only if) the event has started (date+time)."""
    try:
        from .models import Event, Attendance
    except Exception:
        return HttpResponseBadRequest('Models not available')

    mode = 'start'
    # Find the event matching the three start secret numbers
    event = Event.objects.filter(secret_start_a=a, secret_start_b=b, secret_start_c=c).first()
    if not event:
        return render(request, 'frontend/verify_not_found.html', {'a': a, 'b': b, 'c': c, 'mode': mode})

    context = {
        'event': event,
        'mode': mode,
    }

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        if not code:
            messages.error(request, 'Code is required')
            return render(request, 'frontend/verify_start.html', context)

        # Lookup attendance by event + code
        attendance = Attendance.objects.filter(event=event, code=code).first()
        if not attendance:
            messages.error(request, 'Invalid code')
            return render(request, 'frontend/verify_start.html', context)

        # Check if event has started (compare date and time). If start_time is null, compare date only.
        now = datetime.utcnow()
        event_start = datetime.combine(event.start_date, event.start_time or datetime.min.time())

        # If the event has started, record start attendance
        if now >= event_start:
            if not attendance.present_start:
                attendance.present_start = True
                attendance.save(update_fields=['present_start'])
            success_context = {
                'event': event,
                'attendance': attendance,
                'mode': mode,
            }
            return render(request, 'frontend/verify_success.html', success_context)
        else:
            messages.error(request, 'Event has not started yet')
            return render(request, 'frontend/verify_start.html', context)

    return render(request, 'frontend/verify_start.html', context)


def verify_end(request, a, b, c):
    """Same as verify_start but for the end attendance triplet. For now we only
    implement setting present_end later; this route will mirror start logic but
    target the secret_end_* fields."""
    try:
        from .models import Event, Attendance, EventSection
    except Exception:
        return HttpResponseBadRequest('Models not available')

    mode = 'end'
    event = Event.objects.filter(secret_end_a=a, secret_end_b=b, secret_end_c=c).first()
    if not event:
        return render(request, 'frontend/verify_not_found.html', {'a': a, 'b': b, 'c': c, 'mode': mode})

    # Get sections for this event
    sections = EventSection.objects.filter(event=event).order_by('section_code')
    
    context = {
        'event': event,
        'mode': mode,
        'sections': sections,
    }

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        if not code:
            messages.error(request, 'Code is required')
            return render(request, 'frontend/verify_start.html', context)

        attendance = Attendance.objects.filter(event=event, code=code).first()
        if not attendance:
            messages.error(request, 'Invalid code')
            return render(request, 'frontend/verify_start.html', context)
        
        # Add currently selected section IDs to context for pre-checking checkboxes
        context['selected_section_ids'] = list(attendance.sections.values_list('id', flat=True))

        # Determine when end attendance should be allowed. Prefer event end date/time,
        # fall back to start date/time if end info is missing.
        now = datetime.utcnow()
        end_date = event.end_date or event.start_date
        if event.end_time:
            end_time = event.end_time
        elif event.start_time:
            end_time = event.start_time
        else:
            end_time = datetime.min.time()
        event_end = datetime.combine(end_date, end_time)

        if now >= event_end:
            if not attendance.present_end:
                attendance.present_end = True
                attendance.save(update_fields=['present_end'])
            
            # Handle section selection
            selected_section_ids = request.POST.getlist('sections')
            if selected_section_ids:
                attendance.sections.set(selected_section_ids)
            
            success_context = {
                'event': event,
                'attendance': attendance,
                'mode': mode,
            }
            return render(request, 'frontend/verify_success.html', success_context)
        else:
            messages.error(request, 'Event has not ended yet')
            return render(request, 'frontend/verify_start.html', context)

    return render(request, 'frontend/verify_start.html', context)


@login_required
def attendance_qr(request, pk):
    """Member-only page that shows QR links for start and end attendance verification.
    
    The QR codes encode URLs pointing to verify_start and verify_end views with the
    event's three secret numbers as path parameters. For example:
    - Start QR: /events/verify/start/123456/789012/345678/
    - End QR: /events/verify/end/234567/890123/456789/
    
    The QR images are generated using Google Charts API (no external libraries required).
    When attendees scan the QR, they're taken to the verify page where they enter their
    6-digit attendance code to mark their start or end attendance.
    """
    try:
        from .models import Event
    except Exception:
        return HttpResponseBadRequest('Models not available')

    event = get_object_or_404(Event, pk=pk)

    if not (getattr(request.user, 'is_member', False) or request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden('Permission denied')

    # Build verify URLs
    start_ok = event.secret_start_a is not None and event.secret_start_b is not None and event.secret_start_c is not None
    end_ok = event.secret_end_a is not None and event.secret_end_b is not None and event.secret_end_c is not None

    start_url = None
    end_url = None
    if start_ok:
        start_url = request.build_absolute_uri(reverse('verify_start', args=(event.secret_start_a, event.secret_start_b, event.secret_start_c)))
    if end_ok:
        end_url = request.build_absolute_uri(reverse('verify_end', args=(event.secret_end_a, event.secret_end_b, event.secret_end_c)))

    # Generate QR image URLs using QR Server API (Google Charts API is deprecated)
    def qr_img_url(target):
        # URL-encode the target to ensure special characters are properly handled
        encoded_target = quote(target, safe='')
        # Using api.qrserver.com - a free, reliable QR code API
        return f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&data={encoded_target}"

    context = {
        'event': event,
        'start_url': start_url,
        'end_url': end_url,
        'start_qr': qr_img_url(start_url) if start_url else None,
        'end_qr': qr_img_url(end_url) if end_url else None,
    }

    return render(request, 'frontend/attendance_qr.html', context)



@login_required
def create_event(request):
    """Allow portal members to create an event. Only users with is_member True or staff are allowed."""
    # permission check
    if not (request.user.is_authenticated and (request.user.is_staff or getattr(request.user, 'is_member', False))):
        messages.error(request, 'You do not have permission to create events.')
        return redirect('portal')

    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save()
            messages.success(request, f'Event "{event.title}" created successfully.')
            return redirect('events')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = EventForm()

    return render(request, 'frontend/create_event.html', {'form': form})


@login_required
def manage_event_sections(request, pk):
    """Allow members to add/edit/delete professors and sections for an event."""
    if not (getattr(request.user, 'is_member', False) or request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'You do not have permission to manage event sections.')
        return redirect('portal')
    
    try:
        from .models import Event, EventSection
    except Exception:
        messages.error(request, 'Models not available')
        return redirect('portal')
    
    event = get_object_or_404(Event, pk=pk)
    sections = EventSection.objects.filter(event=event).order_by('section_code')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            professor_name = request.POST.get('professor_name', '').strip()
            section_code = request.POST.get('section_code', '').strip()
            
            if professor_name and section_code:
                try:
                    EventSection.objects.create(
                        event=event,
                        professor_name=professor_name,
                        section_code=section_code
                    )
                    messages.success(request, f'Section "{section_code}" added successfully.')
                except Exception as e:
                    messages.error(request, f'Error adding section: {str(e)}')
            else:
                messages.error(request, 'Professor name and section code are required.')
        
        elif action == 'delete':
            section_id = request.POST.get('section_id')
            if section_id:
                try:
                    section = EventSection.objects.get(pk=section_id, event=event)
                    section_name = str(section)
                    section.delete()
                    messages.success(request, f'Section "{section_name}" deleted.')
                except EventSection.DoesNotExist:
                    messages.error(request, 'Section not found.')
                except Exception as e:
                    messages.error(request, f'Error deleting section: {str(e)}')
        
        return redirect('manage_event_sections', pk=event.pk)
    
    return render(request, 'frontend/manage_event_sections.html', {
        'event': event,
        'sections': sections
    })


def image_proxy(request):
    """Simple image proxy for a small set of allowed hosts.

    Use: /image-proxy/?url=<encoded_url>
    This avoids embedding issues when third-party hosts send restrictive CORP/CSP headers.
    """
    url = request.GET.get('url')
    if not url:
        return HttpResponseBadRequest('Missing url')

    parsed = urlparse(url)
    host = parsed.netloc.lower()

    # Whitelist hosts known to be safe for proxying. Adjust as needed.
    allowed_hosts = [
        'photos.fife.usercontent.google.com',
        'lh3.googleusercontent.com',
        'drive.google.com'
    ]

    if host not in allowed_hosts:
        return HttpResponseBadRequest('Host not allowed')

    try:
        resp = requests.get(url, stream=True, timeout=10)
    except Exception as e:
        return HttpResponseBadRequest('Error fetching image')

    if resp.status_code != 200:
        return HttpResponseBadRequest('Upstream returned %s' % resp.status_code)

    content_type = resp.headers.get('Content-Type', 'application/octet-stream')
    data = resp.content

    response = HttpResponse(data, content_type=content_type)
    # We explicitly don't forward CSP/CORP headers from upstream.
    # Allow caching by the browser for a short time; adjust Cache-Control as you prefer.
    response['Cache-Control'] = 'public, max-age=3600'
    return response


@login_required
@require_POST
def delete_event(request, event_id):
    """Allow members or staff to delete an event from the portal."""
    # permission: staff or member
    if not (request.user.is_staff or getattr(request.user, 'is_member', False)):
        return HttpResponseBadRequest('Permission denied')
    # Import Event from local models and fetch
    try:
        from .models import Event
    except Exception:
        return HttpResponseBadRequest('Event model not available')

    event = get_object_or_404(Event, pk=event_id)
    title = event.title
    event.delete()
    messages.success(request, f'Event "{title}" deleted.')
    return redirect('portal')


@login_required
def events_dashboard(request):
    """Events dashboard for members to view all events and attendance data."""
    if not request.user.is_authenticated:
        return HttpResponseRedirect(reverse("login"))
    
    # Check member permission
    if not (getattr(request.user, 'is_member', False) or request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'You do not have permission to access the events dashboard.')
        return redirect('portal')
    
    # Device detection - block mobile and tablet devices
    user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
    if user_agent.is_mobile or user_agent.is_tablet:
        return render(request, 'frontend/desktop_only.html', {
            'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
        })
    
    try:
        from .models import Event, Attendance
        from datetime import datetime
        
        events = Event.objects.all().order_by('-start_date')
        
        # Add attendance stats and status to each event
        events_with_stats = []
        now = datetime.now().date()
        
        for event in events:
            # Determine event status
            if event.end_date < now:
                status = 'past'
            elif event.start_date > now:
                status = 'upcoming'
            else:
                status = 'ongoing'
            
            # Get attendance stats
            attendances = Attendance.objects.filter(event=event)
            total = attendances.count()
            start_count = attendances.filter(present_start=True).count()
            end_count = attendances.filter(present_end=True).count()
            
            event.status = status
            event.total_registered = total
            event.start_attendance = start_count
            event.end_attendance = end_count
            
            events_with_stats.append(event)
        
        return render(request, 'frontend/events_dashboard.html', {
            'events': events_with_stats
        })
    except Exception as e:
        messages.error(request, f'Error loading events dashboard: {str(e)}')
        return redirect('portal')


@login_required
def event_attendance_data(request, pk):
    """API endpoint to fetch attendance data for a specific event."""
    if not (getattr(request.user, 'is_member', False) or request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        from .models import Event, Attendance
        
        event = get_object_or_404(Event, pk=pk)
        filter_type = request.GET.get('filter', 'all')
        
        # Get all attendances for the event
        attendances = Attendance.objects.filter(event=event).order_by('student_name')
        
        # Apply filters
        if filter_type == 'start':
            attendances = attendances.filter(present_start=True)
        elif filter_type == 'end':
            attendances = attendances.filter(present_end=True)
        elif filter_type == 'both':
            attendances = attendances.filter(present_start=True, present_end=True)
        
        # Calculate statistics
        all_attendances = Attendance.objects.filter(event=event)
        total = all_attendances.count()
        start_only = all_attendances.filter(present_start=True, present_end=False).count()
        end_only = all_attendances.filter(present_start=False, present_end=True).count()
        both = all_attendances.filter(present_start=True, present_end=True).count()
        
        # Format attendance data
        attendance_list = []
        for att in attendances:
            # Get sections for this attendance
            sections_list = [f"{s.section_code} ({s.professor_name})" for s in att.sections.all()]
            
            attendance_list.append({
                'student_id': att.student_id,
                'student_name': att.student_name,
                'email': att.email or '',
                'code': att.code,
                'present_start': att.present_start,
                'present_end': att.present_end,
                'sections': ', '.join(sections_list) if sections_list else 'N/A',
            })
        
        return JsonResponse({
            'stats': {
                'total': total,
                'start_only': start_only,
                'end_only': end_only,
                'both': both,
            },
            'attendances': attendance_list
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_event_sections(request, pk):
    """Export event attendance data organized by sections to Excel files (one per section)."""
    if not (getattr(request.user, 'is_member', False) or request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'You do not have permission to export attendance data.')
        return redirect('events_dashboard')
    
    try:
        from .models import Event, EventSection, Attendance
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from io import BytesIO
        import zipfile
        
        event = get_object_or_404(Event, pk=pk)
        sections = EventSection.objects.filter(event=event).order_by('section_code')
        
        if not sections.exists():
            messages.error(request, 'No sections found for this event.')
            return redirect('events_dashboard')
        
        # Create a zip file to hold multiple Excel files (one per section)
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            for section in sections:
                # Create a new workbook for each section
                wb = Workbook()
                ws = wb.active
                ws.title = f"{section.section_code[:25]}"  # Excel sheet name limit is 31 chars
                
                # Title
                ws.merge_cells('A1:E1')
                title_cell = ws['A1']
                title_cell.value = f"{event.title} - Attendance Report"
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30
                
                # Section info
                ws.merge_cells('A2:E2')
                section_cell = ws['A2']
                section_cell.value = f"Section: {section.section_code} - Professor: {section.professor_name}"
                section_cell.font = Font(bold=True, size=12, color="FFFFFF")
                section_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                section_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[2].height = 25
                
                # Event date info
                ws.merge_cells('A3:E3')
                date_cell = ws['A3']
                date_str = f"Event Date: {event.start_date}"
                if event.end_date and event.end_date != event.start_date:
                    date_str += f" to {event.end_date}"
                date_cell.value = date_str
                date_cell.font = Font(italic=True, size=10)
                date_cell.alignment = Alignment(horizontal='center')
                ws.row_dimensions[3].height = 20
                
                # Headers
                headers = ['Student ID', 'Student Name', 'Email', 'Attended Start', 'Attended End']
                header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                ws.row_dimensions[5].height = 25
                
                # Get attendances for this section
                attendances = Attendance.objects.filter(
                    event=event,
                    sections=section
                ).order_by('student_name')
                
                # Data rows
                row_num = 6
                for att in attendances:
                    ws.cell(row=row_num, column=1, value=att.student_id).border = border
                    ws.cell(row=row_num, column=2, value=att.student_name).border = border
                    ws.cell(row=row_num, column=3, value=att.email or 'N/A').border = border
                    
                    # Attended Start - Yes/No
                    start_cell = ws.cell(row=row_num, column=4, value='Yes' if att.present_start else 'No')
                    start_cell.border = border
                    start_cell.alignment = Alignment(horizontal='center')
                    if att.present_start:
                        start_cell.font = Font(color="27AE60", bold=True)
                    else:
                        start_cell.font = Font(color="E74C3C", bold=True)
                    
                    # Attended End - Yes/No
                    end_cell = ws.cell(row=row_num, column=5, value='Yes' if att.present_end else 'No')
                    end_cell.border = border
                    end_cell.alignment = Alignment(horizontal='center')
                    if att.present_end:
                        end_cell.font = Font(color="27AE60", bold=True)
                    else:
                        end_cell.font = Font(color="E74C3C", bold=True)
                    
                    row_num += 1
                
                # Add summary row
                ws.merge_cells(f'A{row_num + 1}:C{row_num + 1}')
                summary_cell = ws.cell(row=row_num + 1, column=1)
                summary_cell.value = f"Total Students: {attendances.count()}"
                summary_cell.font = Font(bold=True, size=11)
                summary_cell.alignment = Alignment(horizontal='left')
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 18
                
                # Save workbook to bytes
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Add to zip file
                safe_section_name = "".join(c for c in section.section_code if c.isalnum() or c in (' ', '-', '_')).strip()
                safe_prof_name = "".join(c for c in section.professor_name if c.isalnum() or c in (' ', '-', '_')).strip()
                filename = f"{safe_section_name}_{safe_prof_name}.xlsx"
                zip_file.writestr(filename, excel_buffer.getvalue())
        
        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        safe_event_name = "".join(c for c in event.title if c.isalnum() or c in (' ', '-', '_')).strip()
        response['Content-Disposition'] = f'attachment; filename="{safe_event_name}_Attendance_by_Section.zip"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting attendance data: {str(e)}')
        return redirect('events_dashboard')


@login_required
def backup_database(request):
    """Download database backup (supports SQLite and PostgreSQL)."""
    # Only superuser or staff can backup database
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'You do not have permission to backup the database.')
        return redirect('portal')
    
    try:
        from django.conf import settings
        from datetime import datetime
        import os
        import subprocess
        import tempfile
        
        db_engine = settings.DATABASES['default']['ENGINE']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # SQLite backup
        if 'sqlite' in db_engine:
            db_path = settings.DATABASES['default']['NAME']
            
            # Check if database file exists
            if not os.path.exists(db_path):
                messages.error(request, 'Database file not found.')
                return redirect('portal')
            
            # Read the database file
            with open(db_path, 'rb') as db_file:
                db_content = db_file.read()
            
            # Create response with database file
            response = HttpResponse(db_content, content_type='application/x-sqlite3')
            response['Content-Disposition'] = f'attachment; filename="dlc_database_backup_{timestamp}.sqlite3"'
            
            messages.success(request, f'Database backup downloaded successfully!')
            return response
        
        # PostgreSQL backup
        elif 'postgresql' in db_engine or 'psycopg2' in db_engine:
            db_config = settings.DATABASES['default']
            db_name = db_config['NAME']
            db_user = db_config['USER']
            db_password = db_config.get('PASSWORD', '')
            db_host = db_config.get('HOST', 'localhost')
            db_port = db_config.get('PORT', '5432')
            
            # Create temporary file for backup
            with tempfile.NamedTemporaryFile(delete=False, suffix='.sql') as tmp_file:
                tmp_path = tmp_file.name
            
            try:
                # Build pg_dump command
                env = os.environ.copy()
                if db_password:
                    env['PGPASSWORD'] = db_password
                
                cmd = [
                    'pg_dump',
                    '-h', db_host,
                    '-p', str(db_port),
                    '-U', db_user,
                    '-F', 'c',  # Custom format (compressed)
                    '-f', tmp_path,
                    db_name
                ]
                
                # Execute pg_dump
                result = subprocess.run(cmd, env=env, capture_output=True, text=True)
                
                if result.returncode != 0:
                    raise Exception(f'pg_dump failed: {result.stderr}')
                
                # Read the backup file
                with open(tmp_path, 'rb') as backup_file:
                    backup_content = backup_file.read()
                
                # Create response
                response = HttpResponse(backup_content, content_type='application/x-postgresql-backup')
                response['Content-Disposition'] = f'attachment; filename="dlc_database_backup_{timestamp}.backup"'
                
                messages.success(request, 'PostgreSQL database backup downloaded successfully!')
                return response
                
            finally:
                # Clean up temporary file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        else:
            messages.error(request, f'Database backup is not supported for {db_engine}.')
            return redirect('portal')
        
    except Exception as e:
        messages.error(request, f'Error backing up database: {str(e)}')
        return redirect('portal')


@login_required
def restore_database(request):
    """Merge/Import data from SQLite file into current database.
    
    Intelligently imports data from uploaded SQLite file:
    - Only adds records that don't already exist
    - Preserves existing data
    - Works with both SQLite and PostgreSQL targets
    """
    import traceback
    import sys
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Only superuser or staff can restore database
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'You do not have permission to import database.')
        return redirect('portal')
    
    # Device detection - block mobile and tablet devices
    try:
        user_agent = parse(request.META.get('HTTP_USER_AGENT', ''))
        if user_agent.is_mobile or user_agent.is_tablet:
            return render(request, 'frontend/desktop_only.html', {
                'device_type': 'mobile device' if user_agent.is_mobile else 'tablet'
            })
    except Exception as e:
        # If device detection fails, continue anyway
        logger.warning(f"Device detection error: {e}")
        pass
    
    if request.method == 'POST':
        tmp_path = None
        sqlite_conn = None
        try:
            import os
            import sqlite3
            import tempfile
            
            logger.info("=== DATABASE IMPORT STARTED ===")
            logger.info(f"User: {request.user.username}")
            logger.info(f"Files in request: {list(request.FILES.keys())}")

            # Get the uploaded file
            uploaded_file = request.FILES.get('database_file')
            
            if not uploaded_file:
                messages.error(request, 'No file uploaded.')
                logger.error("No file uploaded")
                return render(request, 'frontend/restore_database.html')

            logger.info(f"File uploaded: {uploaded_file.name}, size: {uploaded_file.size}")

            # Only accept SQLite files for merge import
            if not uploaded_file.name.endswith(('.sqlite3', '.db', '.sqlite')):
                messages.error(request, 'Please upload a SQLite database file (.sqlite3, .db, or .sqlite).')
                logger.error(f"Invalid file extension: {uploaded_file.name}")
                return render(request, 'frontend/restore_database.html')

            # Save uploaded file to temp location
            logger.info("Saving uploaded file to temp location...")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.sqlite3') as tmp_file:
                for chunk in uploaded_file.chunks():
                    tmp_file.write(chunk)
                tmp_path = tmp_file.name
            
            logger.info(f"Temp file saved to: {tmp_path}")

            # Validate SQLite file
            logger.info("Validating SQLite file...")
            try:
                conn = sqlite3.connect(tmp_path)
                cursor = conn.cursor()
                cursor.execute('PRAGMA schema_version;')
                conn.close()
                logger.info("SQLite file is valid")
            except sqlite3.Error as exc:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                messages.error(request, f'Uploaded file is not a valid SQLite database: {exc}')
                logger.error(f"Invalid SQLite file: {exc}")
                return render(request, 'frontend/restore_database.html')

            # Import all models from main app
            logger.info("Importing models...")
            from main.models import (
                User, Application, Event, Attendance,
                Thread, Reply, DeanList, DeanListStudent, Course, 
                EventSection, ExchangeApplication, PartnerUniversity
            )
            logger.info("Models imported successfully")
            
            messages.info(request, 'Starting data import... This may take a few minutes.')
            messages.warning(request, '⚠️ Large imports may timeout. Keep database files under 1000 records for best results.')
            
            # Connect to uploaded SQLite database
            sqlite_conn = sqlite3.connect(tmp_path)
            sqlite_conn.row_factory = sqlite3.Row
            sqlite_cursor = sqlite_conn.cursor()
            
            stats = {
                'users': {'added': 0, 'skipped': 0},
                'applications': {'added': 0, 'skipped': 0},
                'exchange_apps': {'added': 0, 'skipped': 0},
                'events': {'added': 0, 'skipped': 0},
                'attendance': {'added': 0, 'skipped': 0},
                'threads': {'added': 0, 'skipped': 0},
                'replies': {'added': 0, 'skipped': 0},
                'dean_lists': {'added': 0, 'skipped': 0},
                'dean_students': {'added': 0, 'skipped': 0},
                'courses': {'added': 0, 'skipped': 0},
                'event_sections': {'added': 0, 'skipped': 0},
            }
            
            # Import Users (skip if username exists)
            try:
                from django.utils import timezone
                from datetime import datetime
                
                sqlite_cursor.execute("SELECT * FROM main_user")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert sqlite3.Row to dict
                    username = row.get('username', '')
                    if username and not User.objects.filter(username=username).exists():
                        # Convert naive datetimes to timezone-aware
                        last_login = row.get('last_login')
                        if last_login and isinstance(last_login, str):
                            try:
                                last_login = timezone.make_aware(datetime.fromisoformat(last_login.replace('Z', '+00:00')))
                            except:
                                last_login = None
                        
                        date_joined = row.get('date_joined')
                        if date_joined and isinstance(date_joined, str):
                            try:
                                date_joined = timezone.make_aware(datetime.fromisoformat(date_joined.replace('Z', '+00:00')))
                            except:
                                date_joined = timezone.now()
                        elif not date_joined:
                            date_joined = timezone.now()
                        
                        User.objects.create(
                            username=username,
                            password=row.get('password', ''),
                            last_login=last_login,
                            is_superuser=bool(row.get('is_superuser', 0)),
                            first_name=row.get('first_name', ''),
                            last_name=row.get('last_name', ''),
                            email=row.get('email', ''),
                            is_staff=bool(row.get('is_staff', 0)),
                            is_active=bool(row.get('is_active', 1)),
                            date_joined=date_joined,
                            is_member=bool(row.get('is_member', 0)),
                            role=row.get('role', 'MEMBER'),
                            is_admin=bool(row.get('is_admin', 0)),
                            is_exchange_officer=bool(row.get('is_exchange_officer', 0))
                        )
                        stats['users']['added'] += 1
                    else:
                        stats['users']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'User import issue: {str(e)}')
            
            # Import Events (skip if title + date exists)
            try:
                sqlite_cursor.execute("SELECT * FROM main_event")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    # Get date - might be 'date' or 'event_date' depending on schema
                    event_date = row.get('date') or row.get('event_date')
                    event_title = row.get('title') or row.get('event_title', '')
                    
                    if event_title and event_date:
                        if not Event.objects.filter(title=event_title, date=event_date).exists():
                            Event.objects.create(
                                title=event_title,
                                description=row.get('description', ''),
                                date=event_date,
                                time=row.get('time'),
                                location=row.get('location', ''),
                                created_at=row.get('created_at'),
                                image=row.get('image', '')
                            )
                            stats['events']['added'] += 1
                        else:
                            stats['events']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Event import issue: {str(e)}')
            
            # Import Event Sections
            try:
                sqlite_cursor.execute("SELECT * FROM main_eventsection")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    # Find event by ID from the event we just imported
                    event_id = row.get('event_id')
                    if event_id:
                        # Try to find the event by title since IDs won't match
                        sqlite_cursor.execute("SELECT title, date FROM main_event WHERE id = ?", (event_id,))
                        event_data = sqlite_cursor.fetchone()
                        if event_data:
                            event_data = dict(event_data)  # Convert to dict
                            event = Event.objects.filter(title=event_data['title'], date=event_data['date']).first()
                            if event:
                                section_code = row.get('section_code', '')
                                if not EventSection.objects.filter(event=event, section_code=section_code).exists():
                                    EventSection.objects.create(
                                        event=event,
                                        section_code=section_code,
                                        professor_name=row.get('professor_name', '')
                                    )
                                    stats['event_sections']['added'] += 1
                                else:
                                    stats['event_sections']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Event section import issue: {str(e)}')
            
            # Import Applications (skip if student_id exists for the same academic year)
            try:
                sqlite_cursor.execute("SELECT * FROM main_application")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    student_id = row.get('student_id', '')
                    # Use time_submitted year to avoid duplicates from same year
                    if student_id and not Application.objects.filter(student_id=student_id).exists():
                        Application.objects.create(
                            student_id=student_id,
                            name=row.get('name', ''),
                            email=row.get('email', ''),
                            phone=row.get('phone', ''),
                            passed_credits=row.get('passed_credits', 0),
                            GPA=row.get('GPA', 0.0),
                            major=row.get('major', ''),
                            anything_else=row.get('anything_else', '')
                        )
                        stats['applications']['added'] += 1
                    else:
                        stats['applications']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Application import issue: {str(e)}')
            
            # Import Exchange Applications
            # NOTE: Skipping ExchangeApplications because they have required FileField fields
            # (english_proficiency_document, transcript_document, passport_copy)
            # that cannot be imported from SQLite database
            try:
                messages.info(request, 'Skipping Exchange Applications (requires file uploads)')
                stats['exchange_apps']['skipped'] = 0  # Mark as intentionally skipped
            except Exception as e:
                messages.warning(request, f'Exchange application import issue: {str(e)}')
            
            # Import Attendance
            try:
                sqlite_cursor.execute("SELECT * FROM main_attendance")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    # Find event by looking up the event from source DB and matching by title/date
                    event_id = row.get('event_id')
                    if event_id:
                        sqlite_cursor.execute("SELECT title, date FROM main_event WHERE id = ?", (event_id,))
                        event_data = sqlite_cursor.fetchone()
                        if event_data:
                            event_data = dict(event_data)  # Convert to dict
                            event = Event.objects.filter(title=event_data['title'], date=event_data['date']).first()
                            if event:
                                student_id = row.get('student_id', '')
                                if student_id and not Attendance.objects.filter(event=event, student_id=student_id).exists():
                                    # Find section if it exists
                                    section = None
                                    section_id = row.get('sections_id')
                                    if section_id:
                                        sqlite_cursor.execute("SELECT section_code FROM main_eventsection WHERE id = ?", (section_id,))
                                        section_data = sqlite_cursor.fetchone()
                                        if section_data:
                                            section_data = dict(section_data)  # Convert to dict
                                            section = EventSection.objects.filter(
                                                event=event,
                                                section_code=section_data['section_code']
                                            ).first()
                                    
                                    Attendance.objects.create(
                                        event=event,
                                        student_id=student_id,
                                        student_name=row.get('student_name', ''),
                                        email=row.get('email', ''),
                                        sections=section
                                    )
                                    stats['attendance']['added'] += 1
                                else:
                                    stats['attendance']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Attendance import issue: {str(e)}')
            
            # Import Threads
            try:
                sqlite_cursor.execute("SELECT * FROM main_thread")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    # Find author by looking up in source DB and matching by username
                    author_id = row.get('author_id')
                    if author_id:
                        sqlite_cursor.execute("SELECT username FROM main_user WHERE id = ?", (author_id,))
                        author_data = sqlite_cursor.fetchone()
                        if author_data:
                            author_data = dict(author_data)  # Convert to dict
                            author = User.objects.filter(username=author_data['username']).first()
                            if author:
                                title = row.get('title', '')
                                if title and not Thread.objects.filter(title=title, author=author).exists():
                                    Thread.objects.create(
                                        title=title,
                                        content=row.get('content', ''),
                                        author=author
                                    )
                                    stats['threads']['added'] += 1
                                else:
                                    stats['threads']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Thread import issue: {str(e)}')
            
            # Import Replies
            try:
                sqlite_cursor.execute("SELECT * FROM main_reply")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    # Find thread by looking up in source DB
                    thread_id = row.get('thread_id')
                    author_id = row.get('author_id')
                    
                    if thread_id and author_id:
                        # Get thread title from source
                        sqlite_cursor.execute("SELECT title, author_id FROM main_thread WHERE id = ?", (thread_id,))
                        thread_data = sqlite_cursor.fetchone()
                        
                        # Get author username from source
                        sqlite_cursor.execute("SELECT username FROM main_user WHERE id = ?", (author_id,))
                        author_data = sqlite_cursor.fetchone()
                        
                        if thread_data and author_data:
                            thread_data = dict(thread_data)  # Convert to dict
                            author_data = dict(author_data)  # Convert to dict
                            
                            # Get thread author username
                            sqlite_cursor.execute("SELECT username FROM main_user WHERE id = ?", (thread_data['author_id'],))
                            thread_author_data = sqlite_cursor.fetchone()
                            
                            if thread_author_data:
                                thread_author_data = dict(thread_author_data)  # Convert to dict
                                thread_author = User.objects.filter(username=thread_author_data['username']).first()
                                if thread_author:
                                    thread = Thread.objects.filter(title=thread_data['title'], author=thread_author).first()
                                    author = User.objects.filter(username=author_data['username']).first()
                                    
                                    if thread and author:
                                        Reply.objects.create(
                                            thread=thread,
                                            content=row.get('content', ''),
                                            author=author
                                        )
                                        stats['replies']['added'] += 1
                                    else:
                                        stats['replies']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Reply import issue: {str(e)}')
            
            # Import Dean Lists
            try:
                sqlite_cursor.execute("SELECT * FROM main_deanlist")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    dean_list_name = row.get('name') or row.get('list_name', '')
                    if dean_list_name and not DeanList.objects.filter(name=dean_list_name).exists():
                        DeanList.objects.create(
                            name=dean_list_name,
                            year=row.get('year', ''),
                            semester=row.get('semester', ''),
                            upload_date=row.get('upload_date')
                        )
                        stats['dean_lists']['added'] += 1
                    else:
                        stats['dean_lists']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Dean list import issue: {str(e)}')
            
            # Import Dean List Students
            try:
                sqlite_cursor.execute("SELECT * FROM main_deanliststudent")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    dean_list_id = row.get('dean_list_id')
                    
                    if dean_list_id:
                        # Look up the dean list name from source DB
                        sqlite_cursor.execute("SELECT name FROM main_deanlist WHERE id = ?", (dean_list_id,))
                        dean_list_data = sqlite_cursor.fetchone()
                        
                        if dean_list_data:
                            dean_list_data = dict(dean_list_data)
                            dean_list = DeanList.objects.filter(name=dean_list_data['name']).first()
                            
                            if dean_list:
                                student_id = row.get('student_id', '')
                                if student_id and not DeanListStudent.objects.filter(dean_list=dean_list, student_id=student_id).exists():
                                    DeanListStudent.objects.create(
                                        dean_list=dean_list,
                                        student_id=student_id,
                                        student_name=row.get('student_name', ''),
                                        gpa=row.get('gpa')
                                    )
                                    stats['dean_students']['added'] += 1
                                else:
                                    stats['dean_students']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Dean list student import issue: {str(e)}')
            
            # Import Courses
            try:
                sqlite_cursor.execute("SELECT * FROM main_course")
                for row in sqlite_cursor.fetchall():
                    row = dict(row)  # Convert to dict
                    course_code = row.get('course_code') or row.get('code', '')
                    course_name = row.get('course_name') or row.get('name', '')
                    
                    if course_code and not Course.objects.filter(course_code=course_code).exists():
                        Course.objects.create(
                            course_code=course_code,
                            course_name=course_name,
                            credits=row.get('credits', 3),
                            department=row.get('department', ''),
                            instructor=row.get('instructor', ''),
                            schedule=row.get('schedule', ''),
                            room=row.get('room', ''),
                            semester=row.get('semester', ''),
                            year=row.get('year', '')
                        )
                        stats['courses']['added'] += 1
                    else:
                        stats['courses']['skipped'] += 1
            except Exception as e:
                messages.warning(request, f'Course import issue: {str(e)}')
            
            # Close SQLite connection and clean up
            try:
                if sqlite_conn:
                    sqlite_conn.close()
                    logger.info("SQLite connection closed successfully")
            except Exception as e:
                logger.warning(f"Error closing SQLite connection: {e}")
            
            # Delete temp file
            try:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                    logger.info(f"Temp file deleted: {tmp_path}")
            except Exception as e:
                logger.warning(f"Error deleting temp file: {e}")
            
            logger.info("=== DATABASE IMPORT COMPLETED SUCCESSFULLY ===")
            
            # Build success message
            total_added = sum(s['added'] for s in stats.values())
            total_skipped = sum(s['skipped'] for s in stats.values())
            
            messages.success(request, f'✅ Data import completed! Added {total_added} records, skipped {total_skipped} duplicates.')
            
            # Detailed stats
            details = []
            for model, counts in stats.items():
                if counts['added'] > 0 or counts['skipped'] > 0:
                    details.append(f"{model.replace('_', ' ').title()}: +{counts['added']} new, ~{counts['skipped']} existing")
            
            if details:
                messages.info(request, ' | '.join(details))
            
            return redirect('portal')
            
        except Exception as e:
            import traceback
            from django.conf import settings
            error_trace = traceback.format_exc()
            
            logger.error("=== DATABASE IMPORT ERROR ===")
            logger.error(f"Error type: {type(e).__name__}")
            logger.error(f"Error message: {str(e)}")
            logger.error(f"Full traceback:\n{error_trace}")
            logger.error("=== END ERROR ===")
            
            messages.error(request, f'❌ Error importing database: {type(e).__name__}: {str(e)}')
            
            # Show traceback in development only
            try:
                if settings.DEBUG:
                    for line in error_trace.split('\n')[:10]:  # First 10 lines
                        if line.strip():
                            messages.warning(request, line)
            except:
                pass
        
        finally:
            # ALWAYS close connection and clean up, whether success or error
            # Close SQLite connection if open
            if sqlite_conn:
                try:
                    sqlite_conn.close()
                    logger.info("SQLite connection closed in finally block")
                except Exception as e:
                    logger.warning(f"Error closing connection in finally: {e}")
            
            # Clean up temp file
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                    logger.info(f"Temp file deleted in finally block: {tmp_path}")
                except Exception as cleanup_error:
                    logger.warning(f"Could not delete temp file in finally: {cleanup_error}")
            
        # If we had an error, render the form again
        if 'e' in locals():
            return render(request, 'frontend/restore_database.html')
    
    # GET request - show the upload form
    logger.info("Showing database import form")
    from django.conf import settings
    db_engine = settings.DATABASES['default']['ENGINE']
    context = {
        'current_db': 'PostgreSQL' if ('postgresql' in db_engine or 'psycopg2' in db_engine) else 'SQLite',
        'import_mode': True
    }
    return render(request, 'frontend/restore_database.html', context)
